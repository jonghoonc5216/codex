from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path
from tkinter import (
    BooleanVar,
    Button,
    END,
    LEFT,
    RIGHT,
    BOTH,
    X,
    Y,
    Tk,
    StringVar,
    filedialog,
    messagebox,
)
from tkinter import ttk

import openpyxl


DATA_FILE_NAME = "유사실적back-data-v2.xlsx"
FALLBACK_EXCEL_PATH = Path.home() / "Desktop" / DATA_FILE_NAME
SHEET_NAME = "data1"
DATA_START_ROW = 3


def app_folder() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def find_default_excel_path() -> Path:
    folder = app_folder()
    for candidate in (
        folder / DATA_FILE_NAME,
        folder.parent / DATA_FILE_NAME,
        Path.cwd() / DATA_FILE_NAME,
    ):
        if candidate.exists():
            return candidate
    return FALLBACK_EXCEL_PATH


DEFAULT_EXCEL_PATH = find_default_excel_path()


def to_float(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def to_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if value is None or value == "":
        return None

    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def five_year_cutoff(base_date: dt.date) -> dt.date:
    try:
        return base_date.replace(year=base_date.year - 5)
    except ValueError:
        return base_date.replace(year=base_date.year - 5, day=28)


def five_year_status(completion_date, base_date: dt.date) -> str:
    parsed = to_date(completion_date)
    if not parsed:
        return ""
    return "5년 이내" if parsed >= five_year_cutoff(base_date) else "5년 초과"


def format_value(value):
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        return f"{value:g}"
    if value is None:
        return ""
    return str(value)


def format_result_value(column, value):
    if column in ("총계약금액", "삼안지분금액", "조달청 기준 / 조달청 지분금액"):
        number = to_float(value)
        return f"{number:.1f}" if number is not None else ""
    return format_value(value)


def make_headers(row1, row2):
    headers = []
    current_group = ""
    used = {}
    for index, (top, bottom) in enumerate(zip(row1, row2), start=1):
        if top not in (None, ""):
            current_group = str(top).strip()

        parts = []
        if top not in (None, ""):
            parts.append(str(top).strip())
        elif current_group:
            parts.append(current_group)
        if bottom not in (None, ""):
            parts.append(str(bottom).replace("\n", " ").strip())

        name = " / ".join(parts) if parts else f"컬럼{index}"
        used[name] = used.get(name, 0) + 1
        if used[name] > 1:
            name = f"{name}_{used[name]}"
        headers.append(name)
    return headers


def read_data(path: Path, base_date: dt.date):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"'{SHEET_NAME}' 시트를 찾을 수 없습니다.")

    ws = wb[SHEET_NAME]
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    headers = make_headers(row1, row2)

    rows = []
    for excel_row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if all(cell in (None, "") for cell in excel_row):
            continue
        record = {headers[i]: excel_row[i] if i < len(excel_row) else None for i in range(len(headers))}
        file_five_year = str(record.get("5년여부") or "").strip()
        record["5년여부"] = file_five_year or five_year_status(record.get("준공일"), base_date)
        record["계산_5년여부"] = five_year_status(record.get("준공일"), base_date)
        rows.append(record)
    wb.close()
    return headers, rows


class SimilarResultsApp:
    def __init__(self, root: Tk):
        self.root = root
        self.root.title("유사실적 추출")
        self.root.geometry("1280x760")
        self.root.state("zoomed")

        self.excel_path = StringVar(value=str(DEFAULT_EXCEL_PATH))
        self.base_date = dt.date.today()
        self.base_date_text = StringVar(value=self.base_date.strftime("%Y-%m-%d"))
        self.status_text = StringVar(value="")

        self.use_company_saman = BooleanVar(value=True)
        self.use_company_other = BooleanVar(value=False)
        self.use_five_year_10_to_5 = BooleanVar(value=False)
        self.use_five_year_within_5 = BooleanVar(value=True)
        self.use_completion_amount = BooleanVar(value=False)
        self.completion_amount_min = StringVar(value="200")
        self.use_completion_count = BooleanVar(value=False)
        self.completion_count_min = StringVar(value="")
        self.use_share_amount = BooleanVar(value=False)
        self.share_amount_min = StringVar(value="100")
        self.use_share_count = BooleanVar(value=False)
        self.share_count_min = StringVar(value="")

        self.headers = []
        self.rows = []
        self.filtered_rows = []
        self.selected_result_indexes = set()
        self.excluded_result_keys = set()
        self.five_year_values = []
        self.law_values = []
        self.detail_values = []
        self.participant_values = []
        self.updating_selection_text = False
        self.participant_choice = StringVar(value="전체")

        self.completion_amount_column = StringVar(value="총계약금액")
        self.completion_count_column = StringVar(value="")
        self.share_amount_column = StringVar(value="삼안지분금액")
        self.share_count_column = StringVar(value="")

        self.display_columns = [
            "순번",
            "5년여부",
            "준공일",
            "법령",
            "법령_세부",
            "용역명1",
            "총계약금액",
            "삼안지분금액",
            "준공기준 / 준공1억이상 건수",
            "준공기준 / 준공2억이상 건수",
            "지분기준 / 지분1억이상 건수",
            "지분기준 / 지분2억이상 건수",
            "조달청 기준 / 조달청 지분금액",
            "조달청 기준 / 조달청 지분건수",
        ]

        self.configure_styles()
        self.build_layout()
        self.load_excel()

    def configure_styles(self):
        style = ttk.Style()
        style.configure("Run.TButton", font=("", 12, "bold"), padding=(12, 10))

    def build_layout(self):
        top = ttk.Frame(self.root, padding=10)
        top.pack(fill=X)

        ttk.Label(top, text="엑셀 파일").pack(side=LEFT)
        ttk.Entry(top, textvariable=self.excel_path, width=72).pack(side=LEFT, padx=6)
        ttk.Button(top, text="찾기", command=self.choose_file).pack(side=LEFT)
        ttk.Button(top, text="다시 읽기", command=self.load_excel).pack(side=LEFT, padx=6)
        ttk.Label(top, text="기준일").pack(side=LEFT, padx=(20, 4))
        ttk.Label(top, textvariable=self.base_date_text, width=12).pack(side=LEFT)

        main = ttk.PanedWindow(self.root, orient="horizontal")
        main.pack(fill=BOTH, expand=True, padx=10, pady=(0, 10))

        left_panel = ttk.Frame(main, padding=8, width=340)
        right_panel = ttk.Frame(main, padding=8)
        left_panel.pack_propagate(False)
        left_panel.grid_propagate(False)
        main.add(left_panel, weight=0)
        main.add(right_panel, weight=4)

        self.build_filter_panel(left_panel)
        self.build_result_panel(right_panel)

        bottom = ttk.Frame(self.root, padding=(10, 0, 10, 10))
        bottom.pack(fill=X)
        ttk.Label(bottom, textvariable=self.status_text).pack(side=LEFT)
        ttk.Button(bottom, text="결과 엑셀 저장", command=self.export_results).pack(side=RIGHT)

    def build_filter_panel(self, parent):
        title_frame = ttk.Frame(parent)
        title_frame.pack(fill=X, pady=(0, 8))
        ttk.Label(title_frame, text="조건 선택", font=("", 12, "bold")).pack(side=LEFT)
        ttk.Checkbutton(title_frame, text="삼안", variable=self.use_company_saman).pack(side=LEFT, padx=(42, 2))
        ttk.Checkbutton(title_frame, text="타사", variable=self.use_company_other).pack(side=LEFT, padx=(4, 0))

        five_frame = ttk.LabelFrame(parent, text="5년여부", padding=8)
        five_frame.pack(fill=X, pady=5)
        ttk.Checkbutton(five_frame, text="10년~5년 사이", variable=self.use_five_year_10_to_5).pack(anchor="w")
        ttk.Checkbutton(five_frame, text="5년 이내", variable=self.use_five_year_within_5).pack(anchor="w")

        money_frame = ttk.LabelFrame(parent, text="금액(백만원)", padding=8)
        money_frame.pack(fill=X, pady=5)
        self.add_numeric_filter(
            money_frame,
            "준공금액",
            self.use_completion_amount,
            self.completion_amount_min,
            self.completion_amount_column,
        )
        self.add_numeric_filter(
            money_frame,
            "지분금액",
            self.use_share_amount,
            self.share_amount_min,
            self.share_amount_column,
        )

        law_frame = ttk.LabelFrame(parent, text="법령", padding=8)
        law_frame.pack(fill=X, pady=5)
        self.law_search_vars = [StringVar(value="") for _ in range(3)]
        self.law_entries = []
        law_search_frame = ttk.Frame(law_frame)
        law_search_frame.pack(fill=X, pady=(0, 5))
        for search_var in self.law_search_vars:
            search_var.trace_add("write", lambda *_: self.refresh_law_list())
            entry = ttk.Entry(law_search_frame, textvariable=search_var, width=8)
            entry.pack(side=LEFT, fill=X, expand=True, padx=(0, 3))
            entry.bind(
                "<FocusIn>",
                lambda _event: self.clear_selection_display(self.law_search_vars, self.selected_laws),
            )
            self.law_entries.append(entry)
        law_list_frame = ttk.Frame(law_frame)
        law_list_frame.pack(fill=X)
        self.law_listbox = __import__("tkinter").Listbox(
            law_list_frame, selectmode="extended", exportselection=False, height=8, width=28
        )
        law_scrollbar = ttk.Scrollbar(law_list_frame, orient="vertical", command=self.law_listbox.yview)
        self.law_listbox.configure(yscrollcommand=law_scrollbar.set)
        self.law_listbox.pack(side=LEFT, fill=BOTH, expand=True)
        self.law_listbox.bind("<<ListboxSelect>>", lambda _event: self.show_selected_laws())
        law_scrollbar.pack(side=RIGHT, fill=Y)
        ttk.Button(law_frame, text="선택 해제", command=self.clear_law_selection).pack(fill=X, pady=(5, 0))

        detail_frame = ttk.LabelFrame(parent, text="법령_세부", padding=8)
        detail_frame.pack(fill=X, pady=5)
        self.detail_search_vars = [StringVar(value="") for _ in range(3)]
        self.detail_entries = []
        detail_search_frame = ttk.Frame(detail_frame)
        detail_search_frame.pack(fill=X, pady=(0, 5))
        for search_var in self.detail_search_vars:
            search_var.trace_add("write", lambda *_: self.refresh_detail_list())
            entry = ttk.Entry(detail_search_frame, textvariable=search_var, width=8)
            entry.pack(side=LEFT, fill=X, expand=True, padx=(0, 3))
            entry.bind(
                "<FocusIn>",
                lambda _event: self.clear_selection_display(self.detail_search_vars, self.selected_details),
            )
            self.detail_entries.append(entry)
        list_frame = ttk.Frame(detail_frame)
        list_frame.pack(fill=X)
        self.detail_listbox = __import__("tkinter").Listbox(
            list_frame, selectmode="extended", exportselection=False, height=8, width=28
        )
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.detail_listbox.yview)
        self.detail_listbox.configure(yscrollcommand=scrollbar.set)
        self.detail_listbox.pack(side=LEFT, fill=BOTH, expand=True)
        self.detail_listbox.bind("<<ListboxSelect>>", lambda _event: self.show_selected_details())
        scrollbar.pack(side=RIGHT, fill=Y)
        ttk.Button(detail_frame, text="선택 해제", command=self.clear_detail_selection).pack(fill=X, pady=(5, 0))

        participant_frame = ttk.LabelFrame(parent, text="참여기술자", padding=8)
        participant_frame.pack(fill=X, pady=5)
        participant_list_frame = ttk.Frame(participant_frame)
        participant_list_frame.pack(fill=X)
        self.participant_listbox = __import__("tkinter").Listbox(
            participant_list_frame,
            selectmode="extended",
            exportselection=False,
            height=5,
            width=28,
        )
        participant_scrollbar = ttk.Scrollbar(
            participant_list_frame, orient="vertical", command=self.participant_listbox.yview
        )
        self.participant_listbox.configure(yscrollcommand=participant_scrollbar.set)
        self.participant_listbox.pack(side=LEFT, fill=BOTH, expand=True)
        participant_scrollbar.pack(side=RIGHT, fill=Y)
        ttk.Button(participant_frame, text="선택 해제", command=self.clear_participant_selection).pack(fill=X, pady=(5, 0))

        action_frame = ttk.Frame(parent)
        action_frame.pack(fill=X, pady=(6, 0))
        Button(
            action_frame,
            text="실행",
            command=self.run_filter,
            bg="#0b79d0",
            fg="white",
            activebackground="#095fa3",
            activeforeground="white",
            font=("", 12, "bold"),
            height=2,
            relief="raised",
            cursor="hand2",
        ).pack(side=LEFT, fill=X, expand=True, ipady=7, padx=(0, 4))
        Button(
            action_frame,
            text="전체 선택 해제",
            command=self.clear_all_selections,
            font=("", 12, "bold"),
            height=2,
            relief="raised",
            cursor="hand2",
        ).pack(
            side=LEFT, fill=X, expand=True, ipady=7, padx=(4, 0)
        )

    def add_numeric_filter(self, parent, label, use_var, value_var, column_var):
        row = ttk.Frame(parent)
        row.pack(fill=X, pady=4)
        ttk.Checkbutton(row, text=label, variable=use_var, width=10).pack(side=LEFT)
        ttk.Label(row, text="이상").pack(side=LEFT, padx=(2, 2))
        ttk.Entry(row, textvariable=value_var, width=10).pack(side=LEFT)
        combo = ttk.Combobox(row, textvariable=column_var, state="readonly", width=22)
        combo.pack(side=LEFT, padx=(4, 0), fill=X, expand=True)
        if label == "준공금액":
            self.completion_amount_combo = combo
        elif label == "준공 건수":
            self.completion_count_combo = combo
        elif label == "지분금액":
            self.share_amount_combo = combo
        else:
            self.share_count_combo = combo

    def build_result_panel(self, parent):
        ttk.Label(parent, text="추출 결과", font=("", 12, "bold")).pack(anchor="w", pady=(0, 8))
        self.summary_text = StringVar(value="결과 없음")
        ttk.Label(parent, textvariable=self.summary_text).pack(anchor="w", pady=(0, 6))

        table_frame = ttk.Frame(parent)
        table_frame.pack(fill=BOTH, expand=True)
        self.tree = ttk.Treeview(table_frame, show="headings")
        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=y_scroll.set)
        self.tree.bind("<Button-1>", self.on_result_click)
        self.tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)
        self.tree.bind("<MouseWheel>", self.on_result_mousewheel)

    def choose_file(self):
        current_path = Path(self.excel_path.get())
        initial_dir = current_path.parent if current_path.exists() else app_folder()
        selected = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")),
            initialdir=str(initial_dir),
        )
        if selected:
            self.excel_path.set(selected)
            self.load_excel()

    def load_excel(self):
        path = Path(self.excel_path.get())
        if not path.exists():
            messagebox.showerror("오류", f"파일을 찾을 수 없습니다.\n{path}")
            return

        try:
            self.base_date = dt.date.today()
            self.base_date_text.set(self.base_date.strftime("%Y-%m-%d"))
            self.headers, self.rows = read_data(path, self.base_date)
            self.excluded_result_keys = set()
        except Exception as exc:
            messagebox.showerror("오류", str(exc))
            return

        self.prepare_filter_options()
        self.status_text.set(f"{path.name} 읽기 완료: {len(self.rows)}건")
        self.run_filter()

    def prepare_filter_options(self):
        numeric_columns = []
        for header in ["총계약금액", "삼안지분금액", *self.headers]:
            if header in numeric_columns:
                continue
            if any(to_float(row.get(header)) is not None for row in self.rows):
                numeric_columns.append(header)

        self.completion_amount_combo["values"] = numeric_columns
        self.share_amount_combo["values"] = numeric_columns

        self.set_column_if_present(self.completion_amount_column, numeric_columns, "총계약금액")
        self.set_column_if_present(self.share_amount_column, numeric_columns, "삼안지분금액")
        self.set_column_if_present(self.completion_count_column, numeric_columns, "준공기준 / 준공1억이상 건수")
        self.set_column_if_present(self.share_count_column, numeric_columns, "지분기준 / 지분1억이상 건수")

        self.five_year_values = sorted(
            {str(row.get("5년여부")).strip() for row in self.rows if row.get("5년여부") not in (None, "")}
        )

        self.law_values = sorted(
            {str(row.get("법령")).strip() for row in self.rows if row.get("법령") not in (None, "")}
        )
        self.refresh_law_list()

        self.detail_values = sorted(
            {str(row.get("법령_세부")).strip() for row in self.rows if row.get("법령_세부") not in (None, "")}
        )
        self.refresh_detail_list()

        self.participant_values = [
            header.split(" / ", 1)[1]
            for header in self.headers
            if (
                header.startswith("참여기술자 / ")
                or header.startswith("참여자명단 / ")
            )
            and " / " in header
        ]
        self.refresh_participant_list()

    @staticmethod
    def set_column_if_present(var, columns, preferred):
        if preferred in columns:
            var.set(preferred)
        elif columns:
            var.set(columns[0])

    def refresh_detail_list(self):
        if not hasattr(self, "detail_listbox"):
            return
        if self.updating_selection_text:
            return
        current = self.selected_details()
        queries = self.search_queries(self.detail_search_vars, current)
        self.detail_listbox.delete(0, END)
        shown = [value for value in self.detail_values if self.matches_queries(value, queries)]
        for value in shown:
            self.detail_listbox.insert(END, value)
            if value in current:
                self.detail_listbox.selection_set(END)

    def selected_details(self):
        if not hasattr(self, "detail_listbox"):
            return set()
        return {self.detail_listbox.get(i) for i in self.detail_listbox.curselection()}

    def refresh_law_list(self):
        if not hasattr(self, "law_listbox"):
            return
        if self.updating_selection_text:
            return
        current = self.selected_laws()
        queries = self.search_queries(self.law_search_vars, current)
        self.law_listbox.delete(0, END)
        shown = [value for value in self.law_values if self.matches_queries(value, queries)]
        for value in shown:
            self.law_listbox.insert(END, value)
            if value in current:
                self.law_listbox.selection_set(END)

    def selected_laws(self):
        if not hasattr(self, "law_listbox"):
            return set()
        return {self.law_listbox.get(i) for i in self.law_listbox.curselection()}

    def selected_participants(self):
        if not hasattr(self, "participant_listbox"):
            return set()
        return {self.participant_listbox.get(i) for i in self.participant_listbox.curselection()}

    def refresh_participant_list(self):
        if not hasattr(self, "participant_listbox"):
            return
        current = self.selected_participants()
        self.participant_listbox.delete(0, END)
        for value in self.participant_values:
            self.participant_listbox.insert(END, value)
            if value in current:
                self.participant_listbox.selection_set(END)

    def clear_detail_selection(self):
        self.detail_listbox.selection_clear(0, END)
        self.clear_search_vars(self.detail_search_vars)

    def clear_law_selection(self):
        self.law_listbox.selection_clear(0, END)
        self.clear_search_vars(self.law_search_vars)

    def clear_participant_selection(self):
        self.participant_listbox.selection_clear(0, END)

    def clear_all_selections(self):
        self.use_company_saman.set(True)
        self.use_company_other.set(False)
        self.use_five_year_10_to_5.set(False)
        self.use_five_year_within_5.set(True)
        self.use_completion_amount.set(False)
        self.completion_amount_min.set("200")
        self.use_completion_count.set(False)
        self.completion_count_min.set("")
        self.use_share_amount.set(False)
        self.share_amount_min.set("100")
        self.use_share_count.set(False)
        self.share_count_min.set("")
        self.clear_law_selection()
        self.clear_detail_selection()
        self.clear_participant_selection()
        self.filtered_rows = []
        self.selected_result_indexes = set()
        self.excluded_result_keys = set()
        self.refresh_table()
        self.summary_text.set("결과 없음")
        self.status_text.set("전체 선택 해제 완료")

    def selected_text(self, selected):
        return ", ".join(sorted(selected))

    def search_queries(self, variables, selected):
        selected_text = self.selected_text(selected).lower()
        return [
            variable.get().strip().lower()
            for variable in variables
            if variable.get().strip() and variable.get().strip().lower() != selected_text
        ]

    @staticmethod
    def matches_queries(value, queries):
        if not queries:
            return True
        text = value.lower()
        return any(query in text for query in queries)

    def clear_search_vars(self, variables):
        for variable in variables:
            variable.set("")

    def set_search_from_selection(self, variables, selected):
        self.updating_selection_text = True
        try:
            variables[0].set(self.selected_text(selected))
            for variable in variables[1:]:
                variable.set("")
        finally:
            self.updating_selection_text = False

    def clear_selection_display(self, variables, selected_getter):
        if variables[0].get() == self.selected_text(selected_getter()):
            self.clear_search_vars(variables)

    def show_selected_laws(self):
        self.set_search_from_selection(self.law_search_vars, self.selected_laws())

    def show_selected_details(self):
        self.set_search_from_selection(self.detail_search_vars, self.selected_details())

    def row_has_participant(self, row, participant_name):
        for group_name in ("참여기술자", "참여자명단"):
            value = row.get(f"{group_name} / {participant_name}")
            if value not in (None, ""):
                return True
        return False

    def passes_numeric_filter(self, row, enabled, value_text, column_name):
        if not enabled.get():
            return True
        threshold = to_float(value_text.get())
        if threshold is None:
            raise ValueError("금액/건수 조건에는 숫자를 입력해야 합니다.")
        value = to_float(row.get(column_name.get()))
        return value is not None and value >= threshold

    def run_filter(self):
        selected_companies = set()
        if self.use_company_saman.get():
            selected_companies.add("삼안")
        if self.use_company_other.get():
            selected_companies.add("타사")
        selected_five_years = set()
        if self.use_five_year_10_to_5.get():
            selected_five_years.add("10년~5년 사이")
        if self.use_five_year_within_5.get():
            selected_five_years.add("5년 이내")
        selected_laws = self.selected_laws()
        selected_details = self.selected_details()
        selected_participants = self.selected_participants()
        result = []

        try:
            for row in self.rows:
                if str(row.get("회사명") or "").strip() not in selected_companies:
                    continue
                if selected_five_years and row.get("5년여부") not in selected_five_years:
                    continue
                if not selected_five_years:
                    continue
                if selected_laws and str(row.get("법령")).strip() not in selected_laws:
                    continue
                if selected_details and str(row.get("법령_세부")).strip() not in selected_details:
                    continue
                if selected_participants and not any(
                    self.row_has_participant(row, participant) for participant in selected_participants
                ):
                    continue
                if not self.passes_numeric_filter(
                    row,
                    self.use_completion_amount,
                    self.completion_amount_min,
                    self.completion_amount_column,
                ):
                    continue
                if not self.passes_numeric_filter(
                    row,
                    self.use_completion_count,
                    self.completion_count_min,
                    self.completion_count_column,
                ):
                    continue
                if not self.passes_numeric_filter(
                    row,
                    self.use_share_amount,
                    self.share_amount_min,
                    self.share_amount_column,
                ):
                    continue
                if not self.passes_numeric_filter(row, self.use_share_count, self.share_count_min, self.share_count_column):
                    continue
                result.append(row)
        except ValueError as exc:
            messagebox.showwarning("입력 확인", str(exc))
            return

        self.filtered_rows = result
        self.selected_result_indexes = {
            row_index
            for row_index, row in enumerate(self.filtered_rows)
            if self.result_row_key(row) not in self.excluded_result_keys
        }
        self.refresh_table()
        self.refresh_summary()

    def active_display_columns(self):
        columns = []
        for column in self.display_columns:
            if column in self.headers:
                columns.append(column)
        for column in (
            self.completion_amount_column.get(),
            self.completion_count_column.get(),
            self.share_amount_column.get(),
            self.share_count_column.get(),
        ):
            if column and column not in columns:
                columns.append(column)
        return columns

    def refresh_table(self):
        columns = self.active_display_columns()
        table_columns = ["선택", *columns]
        self.tree["columns"] = table_columns
        heading_labels = self.result_heading_labels()
        compact_columns = {
            "선택": 62,
            "순번": 48,
            "5년여부": 88,
            "준공일": 82,
            "법령": 78,
            "법령_세부": 170,
            "용역명1": 330,
            "총계약금액": 82,
            "삼안지분금액": 82,
            "준공기준 / 준공1억이상 건수": 86,
            "준공기준 / 준공2억이상 건수": 86,
            "지분기준 / 지분1억이상 건수": 86,
            "지분기준 / 지분2억이상 건수": 86,
            "조달청 기준 / 조달청 지분금액": 86,
            "조달청 기준 / 조달청 지분건수": 86,
        }
        centered_keywords = ("금액", "건수")
        right_aligned_columns = {"총계약금액", "삼안지분금액", "조달청 기준 / 조달청 지분금액"}
        for column in table_columns:
            self.tree.heading(column, text=heading_labels.get(column, column))
            if column in compact_columns:
                if column in right_aligned_columns:
                    anchor = "e"
                else:
                    anchor = "center" if column != "용역명1" and column != "법령_세부" else "w"
                self.tree.column(column, width=compact_columns[column], minwidth=40, anchor=anchor, stretch=False)
            elif any(keyword in column for keyword in centered_keywords):
                self.tree.column(column, width=80, minwidth=60, anchor="center", stretch=False)
            else:
                self.tree.column(column, width=120, minwidth=60, anchor="w", stretch=False)
        for item in self.tree.get_children():
            self.tree.delete(item)
        duplicate_tags = self.duplicate_result_tags()
        selected_participants = sorted(self.selected_participants())
        if selected_participants:
            for participant in selected_participants:
                participant_rows = [
                    (row_index, row)
                    for row_index, row in enumerate(self.filtered_rows)
                    if self.row_has_participant(row, participant)
                ]
                for law, rows in self.group_rows_by_law(participant_rows):
                    self.insert_result_rows(columns, rows, duplicate_tags, participant)
                    selected_rows = [row for row_index, row in rows if row_index in self.selected_result_indexes]
                    self.tree.insert(
                        "",
                        END,
                        values=["", *self.total_row_values(columns, selected_rows, f"{law} 소계 {len(selected_rows)}건")],
                        tags=("subtotal",),
                    )
                selected_participant_rows = [
                    row for row_index, row in participant_rows if row_index in self.selected_result_indexes
                ]
                self.tree.insert(
                    "",
                    END,
                    values=["", *self.total_row_values(columns, selected_participant_rows, f"{participant} 합계 {len(selected_participant_rows)}건")],
                    tags=("participant_total",),
                )
        else:
            for law, rows in self.group_rows_by_law():
                self.insert_result_rows(columns, rows, duplicate_tags, law)
                selected_rows = [row for row_index, row in rows if row_index in self.selected_result_indexes]
                self.tree.insert(
                    "",
                    END,
                    values=["", *self.total_row_values(columns, selected_rows, f"{law} 소계 {len(selected_rows)}건")],
                    tags=("subtotal",),
                )
        selected_rows = self.selected_filtered_rows()
        if self.filtered_rows:
            self.tree.insert(
                "",
                END,
                values=["", *self.total_row_values(columns, selected_rows, f"전체 합계 {len(selected_rows)}건")],
                tags=("total",),
            )
            self.tree.tag_configure("subtotal", background="#edf5ff", font=("", 9, "bold"))
            self.tree.tag_configure("participant_total", background="#ffe1e1", font=("", 9, "bold"))
            self.tree.tag_configure("total", background="#f2f2f2", font=("", 9, "bold"))
        self.configure_duplicate_tags()

    def insert_result_rows(self, columns, indexed_rows, duplicate_tags, group_key):
        safe_group = str(group_key).replace(" ", "_").replace("/", "_")
        for row_index, row in indexed_rows:
            checked = "■ 선택" if row_index in self.selected_result_indexes else "□ 해제"
            row_tags = (duplicate_tags[row_index],) if row_index in duplicate_tags else ()
            self.tree.insert(
                "",
                END,
                iid=f"row_{row_index}_{safe_group}",
                values=[checked, *[format_result_value(column, row.get(column)) for column in columns]],
                tags=row_tags,
            )

    def result_heading_labels(self):
        return {
            "준공기준 / 준공1억이상 건수": "준공1억이상 건수",
            "준공기준 / 준공2억이상 건수": "준공2억이상 건수",
            "지분기준 / 지분1억이상 건수": "지분1억 건수",
            "지분기준 / 지분2억이상 건수": "지분2억 건수",
            "총계약금액": "총계약",
            "삼안지분금액": "삼안지분",
            "조달청 기준 / 조달청 지분금액": "조달청 지분금액",
            "조달청 기준 / 조달청 지분건수": "조달청 지분건수",
        }

    def duplicate_result_tags(self):
        seen = {}
        for row_index, row in enumerate(self.filtered_rows):
            if row_index not in self.selected_result_indexes:
                continue
            key = str(row.get("용역명1") or "").strip()
            if not key:
                continue
            seen.setdefault(key, []).append(row_index)
        duplicate_tags = {}
        duplicate_group_index = 0
        for indexes in seen.values():
            if len(indexes) > 1:
                tag_name = f"duplicate_{duplicate_group_index % len(self.duplicate_colors())}"
                for row_index in indexes:
                    duplicate_tags[row_index] = tag_name
                duplicate_group_index += 1
        return duplicate_tags

    @staticmethod
    def duplicate_colors():
        return (
            "#d00000",
            "#0b60d0",
            "#00843d",
            "#9b4d00",
            "#7a1fa2",
            "#c2185b",
            "#006d77",
            "#5d4037",
        )

    def configure_duplicate_tags(self):
        for index, color in enumerate(self.duplicate_colors()):
            self.tree.tag_configure(f"duplicate_{index}", foreground=color)

    def group_rows_by_law(self, indexed_rows=None):
        grouped = {}
        if indexed_rows is None:
            indexed_rows = list(enumerate(self.filtered_rows))
        for row_index, row in indexed_rows:
            law = str(row.get("법령") or "법령 없음").strip()
            grouped.setdefault(law, []).append((row_index, row))
        return [(law, grouped[law]) for law in sorted(grouped)]

    def selected_filtered_rows(self):
        return [
            row
            for row_index, row in enumerate(self.filtered_rows)
            if row_index in self.selected_result_indexes
        ]

    def on_result_click(self, event):
        if self.tree.identify_region(event.x, event.y) != "cell":
            return
        if self.tree.identify_column(event.x) != "#1":
            return
        item_id = self.tree.identify_row(event.y)
        if not item_id.startswith("row_"):
            return
        row_index = int(item_id.split("_")[1])
        row_key = self.result_row_key(self.filtered_rows[row_index])
        if row_index in self.selected_result_indexes:
            self.selected_result_indexes.remove(row_index)
            self.excluded_result_keys.add(row_key)
        else:
            self.selected_result_indexes.add(row_index)
            self.excluded_result_keys.discard(row_key)
        self.refresh_table()
        self.refresh_summary()
        return "break"

    def result_row_key(self, row):
        serial = str(row.get("순번") or "").strip()
        if serial:
            return ("순번", serial)
        return (
            "행",
            str(row.get("준공일") or "").strip(),
            str(row.get("법령") or "").strip(),
            str(row.get("법령_세부") or "").strip(),
            str(row.get("용역명1") or "").strip(),
            str(row.get("총계약금액") or "").strip(),
            str(row.get("삼안지분금액") or "").strip(),
        )

    def on_result_mousewheel(self, event):
        self.tree.yview_scroll(int(-1 * (event.delta / 120)), "units")
        return "break"

    def total_row_values(self, columns, rows, label):
        values = []
        for index, column in enumerate(columns):
            if index == 0:
                values.append(label)
                continue
            if column == "법령":
                values.append(f"계({len(rows)}건)")
                continue
            if any(to_float(row.get(column)) is not None for row in rows):
                total = sum(to_float(row.get(column)) or 0 for row in rows)
                values.append(
                    f"{total:.1f}"
                    if column in ("총계약금액", "삼안지분금액", "조달청 기준 / 조달청 지분금액")
                    else f"{total:g}"
                )
            else:
                values.append("")
        return values

    def refresh_summary(self):
        selected_rows = self.selected_filtered_rows()
        count = len(selected_rows)
        completion_sum = sum(
            to_float(row.get(self.completion_amount_column.get())) or 0 for row in selected_rows
        )
        share_sum = sum(to_float(row.get(self.share_amount_column.get())) or 0 for row in selected_rows)
        details = len({row.get("법령_세부") for row in selected_rows if row.get("법령_세부")})
        self.summary_text.set(
            f"선택 {count}건 / 추출 {len(self.filtered_rows)}건 | 법령_세부 {details}종 | "
            f"{self.completion_amount_column.get()} 합계 {completion_sum:g} | "
            f"{self.share_amount_column.get()} 합계 {share_sum:g}"
        )
        self.status_text.set(f"선택 {count}건 / 추출 {len(self.filtered_rows)}건")

    def export_results(self):
        if not self.filtered_rows:
            messagebox.showinfo("저장", "저장할 결과가 없습니다.")
            return

        default_name = f"유사실적_추출결과_{self.base_date.strftime('%Y%m%d')}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="결과 저장",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=(("Excel files", "*.xlsx"),),
        )
        if not save_path:
            return

        columns = self.active_display_columns()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "추출결과"
        ws.append(["기준일", self.base_date.strftime("%Y-%m-%d")])
        ws.append([self.summary_text.get()])
        ws.append([])
        heading_labels = self.result_heading_labels()
        ws.append([heading_labels.get(column, column) for column in columns])
        selected_participants = sorted(self.selected_participants())
        if selected_participants:
            for participant in selected_participants:
                participant_rows = [
                    (row_index, row)
                    for row_index, row in enumerate(self.filtered_rows)
                    if self.row_has_participant(row, participant)
                ]
                for law, rows in self.group_rows_by_law(participant_rows):
                    selected_rows = [row for row_index, row in rows if row_index in self.selected_result_indexes]
                    if not selected_rows:
                        continue
                    for row in selected_rows:
                        ws.append([format_result_value(column, row.get(column)) for column in columns])
                    ws.append(self.total_row_values(columns, selected_rows, f"{law} 소계 {len(selected_rows)}건"))
                selected_participant_rows = [
                    row for row_index, row in participant_rows if row_index in self.selected_result_indexes
                ]
                if selected_participant_rows:
                    ws.append(
                        self.total_row_values(
                            columns,
                            selected_participant_rows,
                            f"{participant} 합계 {len(selected_participant_rows)}건",
                        )
                    )
        else:
            for law, rows in self.group_rows_by_law():
                selected_rows = [row for row_index, row in rows if row_index in self.selected_result_indexes]
                if not selected_rows:
                    continue
                for row in selected_rows:
                    ws.append([format_result_value(column, row.get(column)) for column in columns])
                ws.append(self.total_row_values(columns, selected_rows, f"{law} 소계 {len(selected_rows)}건"))
        selected_rows = self.selected_filtered_rows()
        ws.append(self.total_row_values(columns, selected_rows, f"전체 합계 {len(selected_rows)}건"))

        for column_cells in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 60)
            ws.column_dimensions[column_cells[0].column_letter].width = width

        wb.save(save_path)
        messagebox.showinfo("저장 완료", f"결과를 저장했습니다.\n{save_path}")


def main():
    root = Tk()
    app = SimilarResultsApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
