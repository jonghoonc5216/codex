import math
import os
import sys
import tempfile
import tkinter as tk
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFont, ImageTk


DEFAULT_EXCEL_NAME = "유사실적back-data-v3.xlsx"


def app_folder() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def find_excel_in_app_folder() -> Optional[Path]:
    folder = app_folder()
    preferred = folder / DEFAULT_EXCEL_NAME
    if preferred.exists():
        return preferred
    matches = sorted(folder.glob("*.xlsx")) + sorted(folder.glob("*.xlsm"))
    return matches[0] if matches else None


@dataclass
class Record:
    key: str
    company: str
    number: str
    year_group: str
    finished_at: str
    over_100m: bool
    category: str
    law: str
    law_detail: str
    service_name: str
    department: str
    manager: str
    share_rate: float
    total_amount: float
    saman_amount: float
    completion_200m_count: float
    completion_100m_count: float
    share_200m_count: float
    share_100m_count: float
    procurement_amount: float
    procurement_count: int
    technicians: set
    search_text: str


def clean_text(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def to_number(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        return 0.0


def format_amount(value: float) -> str:
    if value is None:
        value = 0
    if math.isclose(value, round(value), abs_tol=0.0001):
        return f"{int(round(value)):,}"
    return f"{value:,.1f}"


def format_count(value: float) -> str:
    if value is None:
        value = 0
    if math.isclose(value, round(value), abs_tol=0.0001):
        return f"{int(round(value)):,}"
    return f"{value:,.3f}".rstrip("0").rstrip(".")


def count_value(raw_value, amount: float, share_rate: float, threshold: float = 100) -> float:
    if raw_value not in (None, ""):
        return to_number(raw_value)
    if amount >= threshold:
        return share_rate if share_rate > 0 else 1.0
    return 0.0


def load_records(path: Path):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["data1"] if "data1" in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    first_header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    second_header = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    technician_names = [clean_text(v) for v in second_header[25:] if clean_text(v)]
    technician_columns = {
        25 + offset: clean_text(name)
        for offset, name in enumerate(second_header[25:])
        if clean_text(name)
    }
    technician_fields = {}
    ordered_fields = []
    current_field = ""
    for index in range(25, max(len(first_header), len(second_header))):
        header_field = clean_text(first_header[index] if index < len(first_header) else "")
        if header_field:
            current_field = header_field
            if current_field not in ordered_fields:
                ordered_fields.append(current_field)
        technician = clean_text(second_header[index] if index < len(second_header) else "")
        if technician and current_field:
            technician_fields.setdefault(technician, set()).add(current_field)

    records = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        if not any(row):
            continue

        technicians = set()
        for index, name in technician_columns.items():
            cell_value = clean_text(row[index] if index < len(row) else "")
            if cell_value:
                technicians.add(name)
                technicians.add(cell_value)

        total_amount = to_number(row[12] if len(row) > 12 else 0)
        saman_amount = to_number(row[13] if len(row) > 13 else 0)
        share_rate = to_number(row[11] if len(row) > 11 else 0)
        searchable_parts = [
            clean_text(row[i] if i < len(row) else "")
            for i in [0, 2, 5, 6, 7, 8, 9, 10, 24]
        ]
        searchable_parts.extend(sorted(technicians))

        records.append(
            Record(
                key=f"row-{row_index}",
                company=clean_text(row[0] if len(row) > 0 else ""),
                number=clean_text(row[1] if len(row) > 1 else ""),
                year_group=clean_text(row[2] if len(row) > 2 else ""),
                finished_at=clean_text(row[3] if len(row) > 3 else ""),
                over_100m=to_number(row[4] if len(row) > 4 else 0) >= 1,
                category=clean_text(row[5] if len(row) > 5 else ""),
                law=clean_text(row[6] if len(row) > 6 else ""),
                law_detail=clean_text(row[7] if len(row) > 7 else ""),
                service_name=clean_text(row[8] if len(row) > 8 else ""),
                department=clean_text(row[9] if len(row) > 9 else ""),
                manager=clean_text(row[10] if len(row) > 10 else ""),
                share_rate=share_rate,
                total_amount=total_amount,
                saman_amount=saman_amount,
                completion_200m_count=count_value(row[15] if len(row) > 15 else None, total_amount, share_rate, 200),
                completion_100m_count=count_value(row[14] if len(row) > 14 else None, total_amount, share_rate),
                share_200m_count=count_value(row[19] if len(row) > 19 else None, saman_amount, share_rate, 200),
                share_100m_count=count_value(row[18] if len(row) > 18 else None, saman_amount, share_rate),
                procurement_amount=to_number(row[22] if len(row) > 22 else 0),
                procurement_count=int(to_number(row[23] if len(row) > 23 else 0)),
                technicians=technicians,
                search_text=" ".join(searchable_parts).lower(),
            )
        )

    ordered_technicians = list(dict.fromkeys(technician_names))
    return records, ordered_technicians, ordered_fields, technician_fields


class SearchApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("사책분책 실적 검색기")
        self.geometry("1420x860")
        self.minsize(1120, 720)
        self.after(0, self.show_maximized)

        self.records = []
        self.technicians = []
        self.fields = []
        self.technician_fields = {}
        self.result_sets = [[] for _ in range(5)]
        self.aggregate_items = {}
        self.deselected_records = set()

        self.company_vars = {
            "삼안": tk.BooleanVar(value=True),
            "타사": tk.BooleanVar(value=True),
        }
        self.year_vars = {
            "10년~5년 사이": tk.BooleanVar(value=False),
            "5년 이내": tk.BooleanVar(value=True),
        }
        self.filter_completion_var = tk.BooleanVar(value=False)
        self.filter_share_var = tk.BooleanVar(value=True)
        self.completion_amount_var = tk.StringVar(value="200")
        self.share_amount_var = tk.StringVar(value="100")
        self.technician_var = tk.StringVar(value="선택 안함")
        self.employee_threshold_var = tk.StringVar(value="선택 안함")
        self.keyword_vars = [tk.StringVar() for _ in range(4)]

        self.summary_vars = [
            {
                "count": tk.StringVar(value="0건"),
                "total": tk.StringVar(value="0"),
                "saman": tk.StringVar(value="0"),
                "completion_200_count": tk.StringVar(value="0건"),
                "completion_100_count": tk.StringVar(value="0건"),
                "share_200_count": tk.StringVar(value="0건"),
                "share_100_count": tk.StringVar(value="0건"),
            }
            for _ in range(5)
        ]
        self.overall_card_title_vars = [tk.StringVar(value=f"키워드{i + 1}:-") for i in range(4)]
        self.overall_card_count_vars = [tk.StringVar(value="0") for _ in range(4)]

        self._configure_style()
        self._build_ui()
        self._load_initial_data()

    def show_maximized(self):
        try:
            self.state("zoomed")
        except tk.TclError:
            self.attributes("-zoomed", True)

    def _configure_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        self.configure(background="#ececec")
        style.configure("TFrame", background="#ececec")
        style.configure("Panel.TFrame", background="#ececec", relief="flat")
        style.configure("Filter.TFrame", background="#ececec", relief="flat")
        style.configure("MetricCard.TFrame", background="#f4f4f4", relief="groove", borderwidth=1)
        style.configure("TLabel", background="#ececec", foreground="#111111", font=("Malgun Gothic", 9))
        style.configure("Panel.TLabel", background="#ececec", foreground="#111111", font=("Malgun Gothic", 9))
        style.configure("Title.TLabel", background="#ececec", foreground="#111111", font=("Malgun Gothic", 14, "bold"))
        style.configure("Metric.TLabel", background="#f4f4f4", foreground="#004a88", font=("Malgun Gothic", 12, "bold"))
        style.configure("Small.TLabel", background="#ececec", foreground="#444444", font=("Malgun Gothic", 8))
        style.configure("MetricSmall.TLabel", background="#f4f4f4", foreground="#333333", font=("Malgun Gothic", 8))
        style.configure("TButton", font=("Malgun Gothic", 9), padding=(8, 3))
        style.configure("Accent.TButton", font=("Malgun Gothic", 9, "bold"), padding=(10, 4))
        style.configure("Red.TButton", font=("Malgun Gothic", 11, "bold"), padding=(18, 8), background="#d40000", foreground="#ffffff")
        style.map("Red.TButton", background=[("active", "#a90000"), ("pressed", "#8f0000")], foreground=[("active", "#ffffff"), ("pressed", "#ffffff")])
        style.configure("Black.TButton", font=("Malgun Gothic", 11, "bold"), padding=(18, 8), background="#000000", foreground="#ffffff")
        style.map("Black.TButton", background=[("active", "#333333"), ("pressed", "#111111")], foreground=[("active", "#ffffff"), ("pressed", "#ffffff")])
        style.configure("TCheckbutton", background="#ececec", foreground="#111111", font=("Malgun Gothic", 9))
        style.configure("TLabelframe", background="#ececec", foreground="#111111", bordercolor="#d0d0d0", padding=7)
        style.configure("TLabelframe.Label", background="#ececec", foreground="#111111", font=("Malgun Gothic", 9))
        style.configure("TEntry", fieldbackground="#ffffff", foreground="#000000", padding=2)
        style.configure("TCombobox", fieldbackground="#ffffff", foreground="#000000", padding=2)
        style.configure("TNotebook", background="#ececec", borderwidth=1)
        style.configure("TNotebook.Tab", font=("Malgun Gothic", 9), padding=(12, 5))
        style.configure("Treeview", font=("Malgun Gothic", 9), rowheight=25, background="#ffffff", fieldbackground="#ffffff", foreground="#111111", bordercolor="#b8b8b8", borderwidth=1)
        style.configure("Treeview.Heading", font=("Malgun Gothic", 9, "bold"), background="#dcdcdc", foreground="#111111")
        style.configure("Employee.Treeview", font=("Malgun Gothic", 9), rowheight=25, background="#ffffff", fieldbackground="#ffffff", foreground="#111111")
        style.configure("Employee.Treeview.Heading", font=("Malgun Gothic", 9), background="#ececec", foreground="#d40000")
        style.map("Treeview", background=[("selected", "#0078d7")], foreground=[("selected", "#ffffff")])

    def _build_ui(self):
        root = ttk.Frame(self, padding=10)
        root.pack(fill="both", expand=True)

        header = ttk.Frame(root, style="Panel.TFrame", padding=6)
        header.pack(fill="x", pady=(0, 8))
        ttk.Label(header, text="사책분책 실적 검색기", style="Title.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Button(header, text="엑셀 다시 선택", command=self.select_excel_file).grid(row=0, column=2, padx=(8, 0), sticky="e")
        ttk.Button(header, text="검색", style="Accent.TButton", command=self.search).grid(row=0, column=3, padx=(8, 0), sticky="e")
        ttk.Button(header, text="키워드 초기화", command=self.clear_keywords).grid(row=0, column=4, padx=(8, 0), sticky="e")
        self.path_label = ttk.Label(header, text="", style="Small.TLabel")
        self.path_label.grid(row=1, column=0, columnspan=5, sticky="w", pady=(4, 0))
        header.columnconfigure(1, weight=1)

        filter_frame = ttk.Frame(root, padding=6, style="Filter.TFrame")
        filter_frame.pack(fill="x", pady=(0, 8))

        company_box = ttk.LabelFrame(filter_frame, text="회사")
        company_box.grid(row=0, column=0, sticky="w", padx=(0, 12))
        for idx, (label, var) in enumerate(self.company_vars.items()):
            ttk.Checkbutton(company_box, text=label, variable=var).grid(row=idx, column=0, padx=6, pady=2, sticky="w")

        year_box = ttk.LabelFrame(filter_frame, text="5년여부")
        year_box.grid(row=0, column=1, sticky="w", padx=(0, 12))
        for idx, (label, var) in enumerate(self.year_vars.items()):
            ttk.Checkbutton(year_box, text=label, variable=var).grid(row=idx, column=0, padx=6, pady=2, sticky="w")

        amount_box = ttk.LabelFrame(filter_frame, text="금액 조건")
        amount_box.grid(row=0, column=2, sticky="w", padx=(0, 12))
        ttk.Checkbutton(amount_box, text="준공금액", variable=self.filter_completion_var).grid(row=0, column=0, padx=(4, 4), pady=3, sticky="w")
        ttk.Label(amount_box, text="이상").grid(row=0, column=1, padx=(2, 4), pady=3)
        completion_entry = ttk.Entry(amount_box, textvariable=self.completion_amount_var, width=8)
        completion_entry.grid(row=0, column=2, padx=(0, 8), pady=3)
        completion_entry.bind("<Return>", lambda _event: self.search())
        ttk.Checkbutton(amount_box, text="지분금액", variable=self.filter_share_var).grid(row=1, column=0, padx=(4, 4), pady=3, sticky="w")
        ttk.Label(amount_box, text="이상").grid(row=1, column=1, padx=(2, 4), pady=3)
        share_entry = ttk.Entry(amount_box, textvariable=self.share_amount_var, width=8)
        share_entry.grid(row=1, column=2, padx=(0, 8), pady=3)
        share_entry.bind("<Return>", lambda _event: self.search())

        tech_box = ttk.LabelFrame(filter_frame, text="참여기술자")
        tech_box.grid(row=0, column=3, sticky="w", padx=(0, 12))
        self.tech_combo = ttk.Combobox(
            tech_box,
            textvariable=self.technician_var,
            state="readonly",
            width=12,
            values=["선택 안함"],
        )
        self.tech_combo.grid(row=0, column=0, padx=6, pady=3)
        ttk.Button(
            tech_box,
            text="선택해제",
            command=lambda: self.technician_var.set("선택 안함"),
        ).grid(row=1, column=0, padx=6, pady=3, sticky="ew")

        overall_box = ttk.LabelFrame(filter_frame, text="검색결과(종합)")
        overall_box.grid(row=0, column=4, sticky="nsew")
        for idx in range(4):
            shadow = tk.Frame(overall_box, height=82, bg="#8a8a8a")
            shadow.grid(row=0, column=idx, padx=(8, 8), pady=(3, 6), sticky="nsew")
            shadow.grid_propagate(False)
            card = tk.Frame(shadow, bg="#5b9bd5", highlightthickness=1, highlightbackground="#2f75b5")
            card.place(relx=0, rely=0, relwidth=1, relheight=1, x=-3, y=-3, width=-1, height=-1)
            tk.Label(
                card,
                textvariable=self.overall_card_title_vars[idx],
                bg="#5b9bd5",
                fg="white",
                font=("Malgun Gothic", 11),
                anchor="w",
                justify="left",
            ).pack(fill="x", padx=14, pady=(8, 0))
            count_frame = tk.Frame(card, bg="#5b9bd5")
            count_frame.pack(fill="both", expand=True)
            tk.Label(
                count_frame,
                textvariable=self.overall_card_count_vars[idx],
                bg="#5b9bd5",
                fg="#000000",
                font=("Malgun Gothic", 22, "bold"),
                anchor="e",
            ).pack(side="left", fill="both", expand=True)
            tk.Label(
                count_frame,
                text="건",
                bg="#5b9bd5",
                fg="white",
                font=("Malgun Gothic", 22),
                anchor="w",
            ).pack(side="left", fill="both", expand=True)
        for idx in range(4):
            overall_box.columnconfigure(idx, weight=1, uniform="overall_cards")
        overall_box.rowconfigure(0, weight=1)
        filter_frame.columnconfigure(4, weight=1)

        keyword_frame = ttk.Frame(root, padding=6, style="Panel.TFrame")
        keyword_frame.pack(fill="x", pady=(0, 8))
        for i, var in enumerate(self.keyword_vars):
            ttk.Label(keyword_frame, text=f"키워드 {i + 1}", style="Panel.TLabel").grid(row=0, column=i * 2, padx=(0, 6), sticky="w")
            entry = ttk.Entry(keyword_frame, textvariable=var, width=28)
            entry.grid(row=0, column=i * 2 + 1, padx=(0, 18), sticky="ew")
            entry.bind("<Return>", lambda _event: self.search())
            keyword_frame.columnconfigure(i * 2 + 1, weight=1)

        footer = ttk.Frame(root, height=54)
        footer.pack(side="bottom", fill="x", pady=(8, 0))
        footer.pack_propagate(False)
        tk.Button(
            footer,
            text="엑셀저장(기술자 현황)",
            command=self.export_employee_statistics,
            width=24,
            height=2,
            bg="#d40000",
            fg="#ffffff",
            activebackground="#a90000",
            activeforeground="#ffffff",
            font=("Malgun Gothic", 11, "bold"),
            relief="raised",
            bd=2,
        ).pack(side="left", padx=(0, 8), pady=4)
        tk.Button(
            footer,
            text="엑셀저장(현황+실적)",
            command=self.export_combined_report,
            width=26,
            height=2,
            bg="#000000",
            fg="#ffffff",
            activebackground="#333333",
            activeforeground="#ffffff",
            font=("Malgun Gothic", 11, "bold"),
            relief="raised",
            bd=2,
        ).place(relx=0.5, rely=0.5, anchor="center")
        tk.Button(
            footer,
            text="엑셀 저장",
            command=self.export_to_excel,
            width=18,
            height=2,
            bg="#d40000",
            fg="#ffffff",
            activebackground="#a90000",
            activeforeground="#ffffff",
            font=("Malgun Gothic", 11, "bold"),
            relief="raised",
            bd=2,
        ).pack(side="right", padx=(8, 0), pady=4)
        tk.Label(
            footer,
            text="조경레저부",
            bg="#ececec",
            fg="#d40000",
            font=("Malgun Gothic", 13, "bold"),
        ).pack(side="right", padx=(0, 8), pady=4)

        result_area = ttk.Frame(root)
        result_area.pack(fill="both", expand=True)

        field_panel = ttk.Frame(result_area, padding=(6, 4, 8, 4), style="Panel.TFrame")
        field_panel.pack(side="left", fill="y")
        field_header = ttk.Frame(field_panel, style="Panel.TFrame")
        field_header.pack(fill="x", pady=(0, 8))
        tk.Label(
            field_header,
            text="■ 분야 선택",
            bg="#ececec",
            fg="#ff0000",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).pack(side="left")
        field_list_frame = tk.Frame(field_panel, bg="#ffffff", highlightthickness=1, highlightbackground="#c9c9c9")
        field_list_frame.pack(fill="both", expand=True)
        self.field_listbox = tk.Listbox(
            field_list_frame,
            selectmode="extended",
            exportselection=False,
            width=16,
            height=24,
            activestyle="none",
            font=("Malgun Gothic", 10),
            bg="#ffffff",
            fg="#111111",
            selectbackground="#5b9bd5",
            selectforeground="#ffffff",
            relief="flat",
            borderwidth=0,
        )
        self.field_listbox.pack(side="left", fill="both", expand=True, padx=4, pady=4)
        field_scroll = ttk.Scrollbar(field_list_frame, orient="vertical", command=self.field_listbox.yview)
        field_scroll.pack(side="right", fill="y")
        self.field_listbox.configure(yscrollcommand=field_scroll.set)
        self.field_listbox.bind("<<ListboxSelect>>", self.on_field_selection_changed)
        ttk.Button(
            field_panel,
            text="선택해제",
            command=self.clear_field_selection,
        ).pack(fill="x", pady=(8, 0))

        employee_panel = ttk.Frame(result_area, padding=(6, 4, 10, 4), style="Panel.TFrame")
        employee_panel.pack(side="left", fill="y")
        employee_header = ttk.Frame(employee_panel, style="Panel.TFrame")
        employee_header.pack(fill="x", pady=(0, 8))
        tk.Label(
            employee_header,
            text="■ 기술자별 현황",
            bg="#ececec",
            fg="#ff0000",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).grid(row=0, column=0, sticky="w")
        tk.Label(
            employee_header,
            text="▶필요건수",
            bg="#ececec",
            fg="#111111",
            font=("Malgun Gothic", 9),
            anchor="e",
        ).grid(row=0, column=1, padx=(8, 4), sticky="e")
        self.employee_threshold_combo = ttk.Combobox(
            employee_header,
            textvariable=self.employee_threshold_var,
            state="readonly",
            width=10,
            values=["선택 안함", "1건 이상", "2건 이상", "3건 이상"],
        )
        self.employee_threshold_combo.grid(row=0, column=2, sticky="e")
        self.employee_threshold_combo.bind("<<ComboboxSelected>>", lambda _event: self.refresh_employee_statistics())
        tk.Label(
            employee_header,
            text="(기술자 더블클릭 → 해당 기술자 실적 이동)",
            bg="#ececec",
            fg="#ff0000",
            font=("Malgun Gothic", 8),
            anchor="w",
        ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(2, 0))
        employee_header.columnconfigure(0, weight=1)

        employee_columns = ("name", "keyword1", "keyword2", "keyword3", "keyword4")
        employee_tree_frame = ttk.Frame(employee_panel)
        employee_tree_frame.pack(fill="both", expand=True)
        self.employee_stats_tree = ttk.Treeview(
            employee_tree_frame,
            columns=employee_columns,
            show="headings",
            style="Employee.Treeview",
            height=24,
        )
        employee_headings = {
            "name": "성명",
            "keyword1": "키워드1",
            "keyword2": "키워드2",
            "keyword3": "키워드3",
            "keyword4": "키워드4",
        }
        employee_widths = {"name": 90, "keyword1": 58, "keyword2": 58, "keyword3": 58, "keyword4": 58}
        for column in employee_columns:
            self.employee_stats_tree.heading(column, text=employee_headings[column])
            self.employee_stats_tree.column(column, width=employee_widths[column], minwidth=45, anchor="center", stretch=False)
        self.employee_stats_tree.bind("<Double-1>", self.select_employee_from_statistics)
        self.employee_stats_tree.grid(row=0, column=0, sticky="nsew")
        employee_tree_frame.rowconfigure(0, weight=1)
        employee_tree_frame.columnconfigure(0, weight=1)

        self.notebook = ttk.Notebook(result_area)
        self.notebook.pack(side="left", fill="both", expand=True)
        self.trees = []
        for index in range(4):
            self._add_result_tab(index)
        self._add_result_tab(4, "종합")
        self.mark_aggregate_tab()
        self.notebook_warning = tk.Label(
            self.notebook,
            text="*주의 : 각각의 키워드 검색결과, 용역명이 중복될 수 있음 (종합탭 확인)",
            bg="#ececec",
            fg="#ff0000",
            font=("Malgun Gothic", 10, "bold"),
        )
        self.notebook_warning.place(x=500, y=4)

    def mark_aggregate_tab(self):
        try:
            font_path = Path(r"C:\Windows\Fonts\malgun.ttf")
            font = ImageFont.truetype(str(font_path), 12) if font_path.exists() else ImageFont.load_default()
            self.tab_images = []
            for index, label in enumerate(["키워드 1", "키워드 2", "키워드 3", "키워드 4", "종합"]):
                image = Image.new("RGBA", (72, 20), (0, 0, 0, 0))
                draw = ImageDraw.Draw(image)
                bbox = draw.textbbox((0, 0), label, font=font)
                text_width = bbox[2] - bbox[0]
                text_height = bbox[3] - bbox[1]
                x = (72 - text_width) // 2
                y = (20 - text_height) // 2 - 1
                color = "#d40000" if index == 4 else "#111111"
                draw.text((x, y), label, fill=color, font=font)
                tab_image = ImageTk.PhotoImage(image)
                self.tab_images.append(tab_image)
                self.notebook.tab(index, text="", image=tab_image, compound="center")
        except Exception:
            self.notebook.tab(4, text="종합")

    def _add_result_tab(self, index: int, title: str | None = None):
        tab = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(tab, text=title or f"키워드 {index + 1}")

        metrics = ttk.Frame(tab, padding=(0, 0, 0, 8))
        metrics.pack(fill="x")
        labels = [
            ("검색건수", "count"),
            ("총계약금액(백만원)", "total"),
            ("삼안지분금액(백만원)", "saman"),
            ("준공2억이상 건수", "completion_200_count"),
            ("준공1억이상 건수", "completion_100_count"),
            ("지분2억이상 건수", "share_200_count"),
            ("지분1억이상 건수", "share_100_count"),
        ]
        for col, (label, key) in enumerate(labels):
            box = ttk.Frame(metrics, padding=10, style="MetricCard.TFrame")
            box.grid(row=0, column=col, sticky="ew", padx=(0, 8))
            ttk.Label(box, text=label, style="MetricSmall.TLabel").pack(anchor="w")
            ttk.Label(box, textvariable=self.summary_vars[index][key], style="Metric.TLabel").pack(anchor="w")
            metrics.columnconfigure(col, weight=1)

        bulk_buttons = ttk.Frame(tab, padding=(0, 0, 0, 6))
        bulk_buttons.pack(fill="x")
        ttk.Button(bulk_buttons, text="전체 선택", command=lambda idx=index: self.set_all_selection(idx, True)).pack(side="right", padx=(6, 0))
        ttk.Button(bulk_buttons, text="전체 해제", command=lambda idx=index: self.set_all_selection(idx, False)).pack(side="right")

        columns = (
            "selected",
            "company",
            "year_group",
            "finished_at",
            "law",
            "law_detail",
            "service_name",
            "total_amount",
            "saman_amount",
            "completion_200m_count",
            "completion_100m_count",
            "share_200m_count",
            "share_100m_count",
        )
        tree_frame = ttk.Frame(tab)
        tree_frame.pack(fill="both", expand=True)
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings")

        headings = {
            "selected": "선택/해제",
            "company": "회사",
            "year_group": "5년여부",
            "finished_at": "준공일",
            "law": "법령",
            "law_detail": "법령_세부",
            "service_name": "용역명",
            "total_amount": "총계약금액",
            "saman_amount": "삼안지분금액",
            "completion_200m_count": "준공2억이상 건수",
            "completion_100m_count": "준공1억이상 건수",
            "share_200m_count": "지분2억이상 건수",
            "share_100m_count": "지분1억이상 건수",
        }
        widths = {
            "selected": 64,
            "company": 56,
            "year_group": 92,
            "finished_at": 82,
            "law": 88,
            "law_detail": 112,
            "service_name": 300,
            "total_amount": 88,
            "saman_amount": 92,
            "completion_200m_count": 76,
            "completion_100m_count": 76,
            "share_200m_count": 76,
            "share_100m_count": 76,
        }
        for column in columns:
            tree.heading(column, text=headings[column])
            anchor = "w" if column == "service_name" else "center"
            stretch = column == "service_name"
            tree.column(column, width=widths[column], minwidth=45, anchor=anchor, stretch=stretch)
        tree.bind("<ButtonRelease-1>", self.toggle_tree_selection)

        tree.grid(row=0, column=0, sticky="nsew")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)
        self.trees.append(tree)

    def _load_initial_data(self):
        path = find_excel_in_app_folder()
        if path is not None:
            try:
                self.load_excel(path)
                return
            except Exception as exc:
                messagebox.showerror("엑셀 읽기 오류", f"엑셀 파일을 읽지 못했습니다.\n\n{exc}")
        messagebox.showinfo("엑셀 선택", "프로그램과 같은 폴더에서 엑셀 파일을 찾지 못했습니다.\n실적 엑셀 파일을 선택해 주세요.")
        self.select_excel_file()

    def load_excel(self, path: Path):
        self.records, self.technicians, self.fields, self.technician_fields = load_records(path)
        self.excel_path = path
        self.deselected_records.clear()
        self.path_label.configure(text=f"{path}  |  {len(self.records):,}건")
        self.populate_field_list()
        self.refresh_technician_views()
        if hasattr(self, "employee_stats_tree"):
            self.employee_stats_tree.configure(height=max(1, len(self.visible_technicians())))
        self.search()

    def select_excel_file(self):
        selected = filedialog.askopenfilename(
            title="실적 엑셀 선택",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
            initialdir=str(app_folder()),
        )
        if selected:
            try:
                self.load_excel(Path(selected))
            except Exception as exc:
                messagebox.showerror("엑셀 읽기 오류", str(exc))

    def populate_field_list(self):
        if not hasattr(self, "field_listbox"):
            return
        self.field_listbox.delete(0, tk.END)
        for field in self.fields:
            self.field_listbox.insert(tk.END, field)

    def selected_fields(self):
        if not hasattr(self, "field_listbox"):
            return []
        return [self.field_listbox.get(index) for index in self.field_listbox.curselection()]

    def visible_technicians(self):
        selected_fields = set(self.selected_fields())
        if not selected_fields:
            return self.technicians
        return [
            technician
            for technician in self.technicians
            if self.technician_fields.get(technician, set()) & selected_fields
        ]

    def refresh_technician_views(self):
        visible = self.visible_technicians()
        if hasattr(self, "tech_combo"):
            self.tech_combo.configure(values=["선택 안함"] + visible)
            if self.technician_var.get() != "선택 안함" and self.technician_var.get() not in visible:
                self.technician_var.set("선택 안함")
        if hasattr(self, "employee_stats_tree"):
            self.employee_stats_tree.configure(height=max(1, len(visible)))

    def on_field_selection_changed(self, _event=None):
        self.refresh_technician_views()
        self.search()

    def clear_field_selection(self):
        if hasattr(self, "field_listbox"):
            self.field_listbox.selection_clear(0, tk.END)
        self.refresh_technician_views()
        self.search()

    def base_filter(self, record: Record, include_technician: bool = True) -> bool:
        selected_companies = {name for name, var in self.company_vars.items() if var.get()}
        selected_years = {name for name, var in self.year_vars.items() if var.get()}
        selected_technician = self.technician_var.get()
        completion_threshold = to_number(self.completion_amount_var.get())
        share_threshold = to_number(self.share_amount_var.get())

        if selected_companies and record.company not in selected_companies:
            return False
        if selected_years and record.year_group not in selected_years:
            return False
        if self.filter_completion_var.get() and record.total_amount < completion_threshold:
            return False
        if self.filter_share_var.get() and record.saman_amount < share_threshold:
            return False
        if include_technician and selected_technician and selected_technician != "선택 안함" and selected_technician not in record.technicians:
            return False
        return True

    def keyword_match(self, record: Record, keyword: str) -> bool:
        groups = [
            [word for word in part.strip().lower().split() if word]
            for part in keyword.replace("，", ",").split(",")
            if part.strip()
        ]
        if not groups:
            return False
        return any(all(word in record.search_text for word in group) for group in groups)

    def search(self):
        base_records = [record for record in self.records if self.base_filter(record)]
        for index, keyword_var in enumerate(self.keyword_vars):
            keyword = keyword_var.get()
            matches = [record for record in base_records if self.keyword_match(record, keyword)]
            self.result_sets[index] = matches
            self.populate_tab(index, matches)
        statistics_records = [record for record in self.records if self.base_filter(record, include_technician=False)]
        self.update_employee_statistics(statistics_records)
        self.refresh_aggregate_tab()

    def refresh_employee_statistics(self):
        records = [record for record in self.records if self.base_filter(record, include_technician=False)]
        self.update_employee_statistics(records)

    def update_employee_statistics(self, records):
        if not hasattr(self, "employee_stats_tree"):
            return
        tree = self.employee_stats_tree
        tree.delete(*tree.get_children())
        tree.tag_configure("employee_qualified", background="#fff2cc", foreground="#000000")
        keywords = [keyword_var.get().strip() for keyword_var in self.keyword_vars]
        entered_keyword_indexes = [index for index, keyword in enumerate(keywords) if keyword]
        threshold_text = self.employee_threshold_var.get()
        threshold = (
            1 if threshold_text == "1건 이상"
            else 2 if threshold_text == "2건 이상"
            else 3 if threshold_text == "3건 이상"
            else None
        )

        for row_index, technician in enumerate(self.visible_technicians()):
            counts = []
            numeric_counts = []
            for keyword_index, keyword in enumerate(keywords):
                if not keyword:
                    counts.append("")
                    numeric_counts.append(0)
                    continue
                count = sum(
                    1
                    for record in records
                    if technician in record.technicians
                    and self.keyword_match(record, keyword)
                    and (keyword_index, record.key) not in self.deselected_records
                )
                counts.append(str(count) if count else "")
                numeric_counts.append(count)
            qualified = bool(
                threshold is not None
                and entered_keyword_indexes
                and all(numeric_counts[index] >= threshold for index in entered_keyword_indexes)
            )
            tags = ("employee_qualified",) if qualified else ()
            tree.insert("", "end", iid=f"employee-{row_index}", values=(technician, *counts), tags=tags)

    def select_employee_from_statistics(self, event):
        tree = event.widget
        item_id = tree.identify_row(event.y)
        if not item_id:
            return
        values = tree.item(item_id, "values")
        if not values:
            return
        technician = clean_text(values[0])
        if technician:
            current = self.technician_var.get().strip()
            self.technician_var.set("선택 안함" if current == technician else technician)
            self.search()

    def clear_keywords(self):
        for keyword_var in self.keyword_vars:
            keyword_var.set("")
        self.deselected_records.clear()
        self.aggregate_items.clear()
        for index in range(5):
            self.result_sets[index] = []
            self.populate_tab(index, [])
        statistics_records = [record for record in self.records if self.base_filter(record, include_technician=False)]
        self.update_employee_statistics(statistics_records)

    def populate_tab(self, index: int, records):
        if index == 4:
            self.populate_aggregate_tab()
            return
        tree = self.trees[index]
        tree.delete(*tree.get_children())
        tree.tag_configure("selected_row", background="#d8ebff", foreground="#102033")
        tree.tag_configure("deselected_row", background="#ffffff", foreground="#475569")
        tree.tag_configure("subtotal_row", background="#fff2cc", foreground="#5f4200")

        for record in records:
            selection_key = (index, record.key)
            is_selected = selection_key not in self.deselected_records
            tree.insert(
                "",
                "end",
                iid=record.key,
                tags=("selected_row" if is_selected else "deselected_row",),
                values=(
                    "☑" if is_selected else "☐",
                    record.company,
                    record.year_group,
                    record.finished_at,
                    record.law,
                    record.law_detail,
                    record.service_name,
                    format_amount(record.total_amount),
                    format_amount(record.saman_amount),
                    format_count(record.completion_200m_count),
                    format_count(record.completion_100m_count),
                    format_count(record.share_200m_count),
                    format_count(record.share_100m_count),
                ),
            )
        tree.insert(
            "",
            "end",
            iid=f"keyword-subtotal-{index}",
            tags=("subtotal_row",),
            values=("소계", "0건", "", "", "", "", "", "0", "0", "0", "0", "0", "0"),
        )
        self.update_summary(index)

    def update_keyword_subtotal(self, index: int, records):
        if index >= 4:
            return
        tree = self.trees[index]
        subtotal_id = f"keyword-subtotal-{index}"
        if not tree.exists(subtotal_id):
            return
        tree.item(
            subtotal_id,
            values=(
                "소계",
                f"{len(records)}건",
                "",
                "",
                "",
                "",
                "",
                format_amount(sum(record.total_amount for record in records)),
                format_amount(sum(record.saman_amount for record in records)),
                format_count(sum(record.completion_200m_count for record in records)),
                format_count(sum(record.completion_100m_count for record in records)),
                format_count(sum(record.share_200m_count for record in records)),
                format_count(sum(record.share_100m_count for record in records)),
            ),
            tags=("subtotal_row",),
        )

    def aggregate_iid(self, keyword_index: int, record_key: str) -> str:
        return f"k{keyword_index}:{record_key}"

    def refresh_aggregate_tab(self):
        if len(self.trees) < 5:
            return
        self.populate_aggregate_tab()

    def populate_aggregate_tab(self):
        tree = self.trees[4]
        tree.delete(*tree.get_children())
        tree.tag_configure("selected_row", background="#d8ebff", foreground="#102033")
        tree.tag_configure("deselected_row", background="#ffffff", foreground="#475569")
        tree.tag_configure("aggregate_duplicate_row", background="#d8ebff", foreground="#d40000")
        tree.tag_configure("section_row", background="#1f4e78", foreground="#ffffff")
        tree.tag_configure("subtotal_row", background="#fff2cc", foreground="#5f4200")
        self.aggregate_items.clear()

        for keyword_index in range(4):
            records = self.result_sets[keyword_index]
            if not records:
                continue
            keyword = self.keyword_vars[keyword_index].get().strip() or "-"
            section_id = f"section-{keyword_index}"
            tree.insert(
                "",
                "end",
                iid=section_id,
                tags=("section_row",),
                values=(
                    f"키워드 {keyword_index + 1}",
                    "",
                    "",
                    "",
                    "",
                    "",
                    f"키워드 {keyword_index + 1}: {keyword}",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ),
            )

            for record in records:
                item_id = self.aggregate_iid(keyword_index, record.key)
                is_selected = (keyword_index, record.key) not in self.deselected_records
                self.aggregate_items[item_id] = (keyword_index, record)
                tree.insert(
                    "",
                    "end",
                    iid=item_id,
                    tags=("selected_row" if is_selected else "deselected_row",),
                    values=(
                        "☑" if is_selected else "☐",
                        record.company,
                        record.year_group,
                        record.finished_at,
                        record.law,
                        record.law_detail,
                        record.service_name,
                        format_amount(record.total_amount),
                        format_amount(record.saman_amount),
                        format_count(record.completion_200m_count),
                        format_count(record.completion_100m_count),
                        format_count(record.share_200m_count),
                        format_count(record.share_100m_count),
                    ),
                )
            subtotal_id = f"subtotal-{keyword_index}"
            tree.insert(
                "",
                "end",
                iid=subtotal_id,
                tags=("subtotal_row",),
                values=(
                    "소계",
                    f"{len(records)}건",
                    "",
                    "",
                    "",
                    "",
                    "",
                    format_amount(sum(record.total_amount for record in records)),
                    format_amount(sum(record.saman_amount for record in records)),
                    format_count(sum(record.completion_200m_count for record in records)),
                    format_count(sum(record.completion_100m_count for record in records)),
                    format_count(sum(record.share_200m_count for record in records)),
                    format_count(sum(record.share_100m_count for record in records)),
                ),
            )
        self.update_aggregate_subtotals()
        self.update_summary(4)

    def aggregate_checked_records(self, keyword_index: int):
        tree = self.trees[4]
        records = []
        for item_id, (source_index, record) in self.aggregate_items.items():
            if source_index != keyword_index or not tree.exists(item_id):
                continue
            values = tree.item(item_id, "values")
            if values and values[0] == "☑":
                records.append(record)
        return records

    def update_aggregate_subtotals(self):
        if len(self.trees) < 5:
            return
        tree = self.trees[4]
        for keyword_index in range(4):
            subtotal_id = f"subtotal-{keyword_index}"
            if not tree.exists(subtotal_id):
                continue
            records = self.aggregate_checked_records(keyword_index)
            tree.item(
                subtotal_id,
                values=(
                    "소계",
                    f"{len(records)}건",
                    "",
                    "",
                    "",
                    "",
                    "",
                    format_amount(sum(record.total_amount for record in records)),
                    format_amount(sum(record.saman_amount for record in records)),
                    format_count(sum(record.completion_200m_count for record in records)),
                    format_count(sum(record.completion_100m_count for record in records)),
                    format_count(sum(record.share_200m_count for record in records)),
                    format_count(sum(record.share_100m_count for record in records)),
                ),
                tags=("subtotal_row",),
            )
        self.update_aggregate_duplicate_styles()

    def duplicate_service_key(self, service_name: str) -> str:
        return " ".join(clean_text(service_name).split()).lower()

    def update_aggregate_duplicate_styles(self):
        if len(self.trees) < 5:
            return
        tree = self.trees[4]
        service_counts = {}
        for item_id, (_source_index, record) in self.aggregate_items.items():
            if not tree.exists(item_id):
                continue
            values = tree.item(item_id, "values")
            if values and values[0] == "☑":
                key = self.duplicate_service_key(record.service_name)
                if key:
                    service_counts[key] = service_counts.get(key, 0) + 1

        for item_id, (_source_index, record) in self.aggregate_items.items():
            if not tree.exists(item_id):
                continue
            values = tree.item(item_id, "values")
            if not values:
                continue
            if values[0] != "☑":
                tree.item(item_id, tags=("deselected_row",))
                continue
            key = self.duplicate_service_key(record.service_name)
            if key and service_counts.get(key, 0) > 1:
                tree.item(item_id, tags=("aggregate_duplicate_row",))
            else:
                tree.item(item_id, tags=("selected_row",))

    def selected_records_for_tab(self, index: int):
        if index == 4 and index < len(self.trees):
            tree = self.trees[index]
            records = []
            for item_id in tree.get_children():
                values = tree.item(item_id, "values")
                if values and values[0] == "☑" and item_id in self.aggregate_items:
                    records.append(self.aggregate_items[item_id][1])
            return records
        if index < len(self.trees):
            tree = self.trees[index]
            children = tree.get_children()
            if children:
                checked_keys = {
                    item_id
                    for item_id in children
                    if tree.item(item_id, "values") and tree.item(item_id, "values")[0] == "☑"
                }
                return [record for record in self.result_sets[index] if record.key in checked_keys]
        return [record for record in self.result_sets[index] if (index, record.key) not in self.deselected_records]

    def update_summary(self, index: int):
        records = self.selected_records_for_tab(index)
        total_amount = sum(record.total_amount for record in records)
        saman_amount = sum(record.saman_amount for record in records)
        completion_200_count = sum(record.completion_200m_count for record in records)
        completion_100_count = sum(record.completion_100m_count for record in records)
        share_200_count = sum(record.share_200m_count for record in records)
        share_100_count = sum(record.share_100m_count for record in records)
        self.summary_vars[index]["count"].set(f"{len(records):,}건")
        self.summary_vars[index]["total"].set(format_amount(total_amount))
        self.summary_vars[index]["saman"].set(format_amount(saman_amount))
        self.summary_vars[index]["completion_200_count"].set(f"{format_count(completion_200_count)}건")
        self.summary_vars[index]["completion_100_count"].set(f"{format_count(completion_100_count)}건")
        self.summary_vars[index]["share_200_count"].set(f"{format_count(share_200_count)}건")
        self.summary_vars[index]["share_100_count"].set(f"{format_count(share_100_count)}건")
        if index < 4:
            self.update_keyword_subtotal(index, records)
            self.update_overall_summary(index, len(records))

    def update_overall_summary(self, index: int, count: int):
        keyword = self.keyword_vars[index].get().strip()
        keyword_text = keyword if keyword else "-"
        self.overall_card_title_vars[index].set(f"키워드{index + 1}:{keyword_text}")
        self.overall_card_count_vars[index].set(f"{count:,}")

    def toggle_tree_selection(self, event):
        tree = event.widget
        if tree.identify_column(event.x) != "#1":
            return
        item_id = tree.identify_row(event.y)
        if not item_id:
            return
        current_index = self.trees.index(tree)
        if item_id.startswith("keyword-subtotal-") or item_id.startswith("subtotal-") or item_id.startswith("section-"):
            return
        if current_index == 4:
            self.toggle_aggregate_selection(item_id)
            return
        selection_key = (current_index, item_id)
        if selection_key in self.deselected_records:
            self.deselected_records.remove(selection_key)
            mark = "☑"
            tags = ("selected_row",)
        else:
            self.deselected_records.add(selection_key)
            mark = "☐"
            tags = ("deselected_row",)
        values = list(tree.item(item_id, "values"))
        values[0] = mark
        tree.item(item_id, values=values, tags=tags)
        self.update_summary(current_index)
        self.refresh_aggregate_tab()
        self.refresh_employee_statistics()

    def set_keyword_item_state(self, keyword_index: int, record_key: str, selected: bool):
        selection_key = (keyword_index, record_key)
        if selected:
            self.deselected_records.discard(selection_key)
            mark = "☑"
            tags = ("selected_row",)
        else:
            self.deselected_records.add(selection_key)
            mark = "☐"
            tags = ("deselected_row",)

        if keyword_index < len(self.trees):
            tree = self.trees[keyword_index]
            if tree.exists(record_key):
                values = list(tree.item(record_key, "values"))
                if values:
                    values[0] = mark
                    tree.item(record_key, values=values, tags=tags)
        self.update_summary(keyword_index)

    def toggle_aggregate_selection(self, item_id: str):
        if item_id not in self.aggregate_items:
            return
        keyword_index, record = self.aggregate_items[item_id]
        tree = self.trees[4]
        values = list(tree.item(item_id, "values"))
        if not values:
            return
        selected = values[0] != "☑"
        values[0] = "☑" if selected else "☐"
        tree.item(item_id, values=values, tags=("selected_row" if selected else "deselected_row",))
        self.set_keyword_item_state(keyword_index, record.key, selected)
        self.update_aggregate_subtotals()
        self.update_summary(4)
        self.refresh_employee_statistics()

    def set_all_selection(self, index: int, selected: bool):
        if index >= len(self.trees):
            return
        tree = self.trees[index]
        if index == 4:
            for item_id in tree.get_children():
                if item_id not in self.aggregate_items:
                    continue
                keyword_index, record = self.aggregate_items[item_id]
                values = list(tree.item(item_id, "values"))
                if not values:
                    continue
                values[0] = "☑" if selected else "☐"
                tree.item(item_id, values=values, tags=("selected_row" if selected else "deselected_row",))
                self.set_keyword_item_state(keyword_index, record.key, selected)
            self.update_aggregate_subtotals()
            self.update_summary(4)
            self.refresh_employee_statistics()
            return
        for item_id in tree.get_children():
            if item_id.startswith("keyword-subtotal-") or item_id.startswith("subtotal-") or item_id.startswith("section-"):
                continue
            selection_key = (index, item_id)
            values = list(tree.item(item_id, "values"))
            if not values:
                continue
            if selected:
                self.deselected_records.discard(selection_key)
                values[0] = "☑"
                tags = ("selected_row",)
            else:
                self.deselected_records.add(selection_key)
                values[0] = "☐"
                tags = ("deselected_row",)
            tree.item(item_id, values=values, tags=tags)
        self.update_summary(index)
        self.refresh_aggregate_tab()
        self.refresh_employee_statistics()

    def export_employee_statistics(self):
        if not hasattr(self, "employee_stats_tree"):
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = Path(tempfile.gettempdir()) / f"사책분책_기술자_현황_미리보기_{timestamp}.xlsx"

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "기술자 현황"

        title_fill = PatternFill("solid", fgColor="1F4E78")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        qualified_fill = PatternFill("solid", fgColor="FFF2CC")
        thin = Side(style="thin", color="B7C9DA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        sheet.merge_cells("A1:E1")
        title_cell = sheet["A1"]
        title_cell.value = "전직원 기술자별 키워드 현황"
        title_cell.fill = title_fill
        title_cell.font = Font(color="FFFFFF", bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        keywords = [keyword_var.get().strip() for keyword_var in self.keyword_vars]
        headers = ["성명"] + [
            f"키워드 {index + 1}\n{keyword}" if keyword else f"키워드 {index + 1}"
            for index, keyword in enumerate(keywords)
        ]
        for column, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=column, value=header)
            cell.fill = header_fill
            cell.font = Font(color="D40000", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        row_number = 3
        for item_id in self.employee_stats_tree.get_children():
            values = self.employee_stats_tree.item(item_id, "values")
            tags = self.employee_stats_tree.item(item_id, "tags")
            highlighted = "employee_qualified" in tags
            for column, value in enumerate(values, 1):
                cell = sheet.cell(row=row_number, column=column, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(color="000000")
                if highlighted:
                    cell.fill = qualified_fill
            row_number += 1

        sheet.column_dimensions["A"].width = 14
        for column in ("B", "C", "D", "E"):
            sheet.column_dimensions[column].width = 18
        sheet.row_dimensions[1].height = 24
        sheet.row_dimensions[2].height = 34
        sheet.freeze_panes = "A3"
        workbook.save(path)

        try:
            os.startfile(path)
        except OSError as exc:
            messagebox.showinfo(
                "기술자 현황 미리보기",
                f"엑셀 파일을 만들었지만 자동으로 열지 못했습니다.\n\n{path}\n\n{exc}",
            )

    def export_combined_report(self):
        keyword_indexes = [index for index, var in enumerate(self.keyword_vars) if var.get().strip()]
        if not keyword_indexes:
            messagebox.showinfo("엑셀 저장", "저장할 키워드가 없습니다.")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = Path(tempfile.gettempdir()) / f"사책분책_현황과실적_미리보기_{timestamp}.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "현황+실적"
        sheet.sheet_view.showGridLines = True

        dark_fill = PatternFill("solid", fgColor="1F4E78")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        selected_fill = PatternFill("solid", fgColor="D8EBFF")
        subtotal_fill = PatternFill("solid", fgColor="FFF2CC")
        qualified_fill = PatternFill("solid", fgColor="FFF2CC")
        thin = Side(style="thin", color="9FBAD0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Left: employee statistics
        sheet.merge_cells("A1:E1")
        employee_title = sheet["A1"]
        employee_title.value = "전직원 기술자별 키워드 현황"
        employee_title.fill = dark_fill
        employee_title.font = Font(color="FFFFFF", bold=True)
        employee_title.alignment = Alignment(horizontal="center", vertical="center")

        keyword_texts = [var.get().strip() for var in self.keyword_vars]
        employee_headers = ["성명"] + [
            f"키워드 {index + 1}\n{keyword}" if keyword else f"키워드 {index + 1}"
            for index, keyword in enumerate(keyword_texts)
        ]
        for column, header in enumerate(employee_headers, 1):
            cell = sheet.cell(row=2, column=column, value=header)
            cell.fill = header_fill
            cell.font = Font(color="D40000", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        employee_row = 3
        for item_id in self.employee_stats_tree.get_children():
            values = self.employee_stats_tree.item(item_id, "values")
            highlighted = "employee_qualified" in self.employee_stats_tree.item(item_id, "tags")
            for column, value in enumerate(values, 1):
                cell = sheet.cell(row=employee_row, column=column, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(color="000000")
                if highlighted:
                    cell.fill = qualified_fill
            employee_row += 1

        # Right: selected performance records
        result_headers = [
            "선택", "회사", "5년여부", "준공일", "법령", "법령_세부", "용역명",
            "총계약금액", "삼안지분금액", "준공2억이상\n건수", "준공1억이상\n건수",
            "지분2억이상\n건수", "지분1억이상\n건수",
        ]
        start_column = 7
        result_row = 1
        for keyword_index in keyword_indexes:
            keyword = keyword_texts[keyword_index]
            records = self.selected_records_for_tab(keyword_index)
            end_column = start_column + len(result_headers) - 1
            sheet.merge_cells(
                start_row=result_row,
                start_column=start_column,
                end_row=result_row,
                end_column=end_column,
            )
            title_cell = sheet.cell(
                row=result_row,
                column=start_column,
                value=f"키워드 {keyword_index + 1}: {keyword}",
            )
            title_cell.fill = dark_fill
            title_cell.font = Font(color="FFFFFF", bold=True)
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            result_row += 1

            for offset, header in enumerate(result_headers):
                cell = sheet.cell(row=result_row, column=start_column + offset, value=header)
                cell.fill = header_fill
                cell.font = Font(color="1F3A5A", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            result_row += 1

            for record in records:
                values = [
                    "☑", record.company, record.year_group, record.finished_at, record.law,
                    record.law_detail, record.service_name, record.total_amount, record.saman_amount,
                    record.completion_200m_count, record.completion_100m_count,
                    record.share_200m_count, record.share_100m_count,
                ]
                for offset, value in enumerate(values):
                    column = start_column + offset
                    cell = sheet.cell(row=result_row, column=column, value=value)
                    cell.fill = selected_fill
                    cell.border = border
                    cell.alignment = Alignment(
                        horizontal="left" if offset == 6 else "center",
                        vertical="center",
                        wrap_text=offset == 6,
                    )
                    if offset >= 7:
                        cell.number_format = "#,##0.###"
                result_row += 1

            subtotal_values = [
                "소계", f"{len(records)}건", "", "", "", "", "",
                sum(record.total_amount for record in records),
                sum(record.saman_amount for record in records),
                sum(record.completion_200m_count for record in records),
                sum(record.completion_100m_count for record in records),
                sum(record.share_200m_count for record in records),
                sum(record.share_100m_count for record in records),
            ]
            for offset, value in enumerate(subtotal_values):
                cell = sheet.cell(row=result_row, column=start_column + offset, value=value)
                cell.fill = subtotal_fill
                cell.font = Font(color="5F4200", bold=True)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if offset >= 7:
                    cell.number_format = "#,##0.###"
            result_row += 2

        widths = {
            "A": 12, "B": 10, "C": 10, "D": 10, "E": 10, "F": 3,
            "G": 7, "H": 9, "I": 12, "J": 11, "K": 14, "L": 18, "M": 48,
            "N": 13, "O": 13, "P": 12, "Q": 12, "R": 12, "S": 12,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        last_row = max(employee_row - 1, result_row - 1, 45)
        sheet.print_area = f"A1:S{last_row}"
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_margins.left = 0.2
        sheet.page_margins.right = 0.2
        sheet.page_margins.top = 0.3
        sheet.page_margins.bottom = 0.3
        sheet.sheet_view.zoomScale = 60
        sheet.freeze_panes = "G3"
        workbook.save(path)

        try:
            os.startfile(path)
        except OSError as exc:
            messagebox.showinfo(
                "현황+실적 미리보기",
                f"엑셀 파일을 만들었지만 자동으로 열지 못했습니다.\n\n{path}\n\n{exc}",
            )

    def export_to_excel(self):
        keyword_indexes = [i for i, var in enumerate(self.keyword_vars) if var.get().strip()]
        if not keyword_indexes:
            messagebox.showinfo("엑셀 저장", "저장할 키워드가 없습니다.")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = Path(tempfile.gettempdir()) / f"사책분책_실적_검색결과_미리보기_{timestamp}.xlsx"

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "검색결과"

        title_fill = PatternFill("solid", fgColor="1F4E78")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        subtotal_fill = PatternFill("solid", fgColor="FFF2CC")
        selected_fill = PatternFill("solid", fgColor="D8EBFF")
        thin = Side(style="thin", color="B7C9DA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        headers = [
            "선택",
            "회사",
            "5년여부",
            "준공일",
            "법령",
            "법령_세부",
            "용역명",
            "총계약금액",
            "삼안지분금액",
            "준공2억이상\n건수",
            "준공1억이상\n건수",
            "지분2억이상\n건수",
            "지분1억이상\n건수",
        ]

        row_number = 1
        for index in keyword_indexes:
            keyword = self.keyword_vars[index].get().strip()
            records = self.selected_records_for_tab(index)
            sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=len(headers))
            cell = sheet.cell(row=row_number, column=1, value=f"키워드 {index + 1}: {keyword}")
            cell.fill = title_fill
            cell.font = Font(color="FFFFFF", bold=True, size=12)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            row_number += 1

            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row_number, column=col, value=header)
                cell.fill = header_fill
                cell.font = Font(bold=True, color="1F3A5A")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            row_number += 1

            for record in records:
                values = [
                    "☑",
                    record.company,
                    record.year_group,
                    record.finished_at,
                    record.law,
                    record.law_detail,
                    record.service_name,
                    record.total_amount,
                    record.saman_amount,
                    record.completion_200m_count,
                    record.completion_100m_count,
                    record.share_200m_count,
                    record.share_100m_count,
                ]
                for col, value in enumerate(values, 1):
                    cell = sheet.cell(row=row_number, column=col, value=value)
                    cell.fill = selected_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left" if col == 7 else "center", vertical="center", wrap_text=(col == 7))
                    if col in (8, 9, 10, 11, 12, 13):
                        cell.number_format = "#,##0.###"
                row_number += 1

            totals = self.selected_records_for_tab(index)
            subtotal_values = [
                "소계",
                f"{len(totals)}건",
                "",
                "",
                "",
                "",
                "",
                sum(record.total_amount for record in totals),
                sum(record.saman_amount for record in totals),
                sum(record.completion_200m_count for record in totals),
                sum(record.completion_100m_count for record in totals),
                sum(record.share_200m_count for record in totals),
                sum(record.share_100m_count for record in totals),
            ]
            for col, value in enumerate(subtotal_values, 1):
                cell = sheet.cell(row=row_number, column=col, value=value)
                cell.fill = subtotal_fill
                cell.font = Font(bold=True, color="5F4200")
                cell.border = border
                cell.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
                if col in (8, 9, 10, 11, 12, 13):
                    cell.number_format = "#,##0.###"
            row_number += 2

        widths = [8, 10, 15, 12, 16, 20, 58, 15, 15, 14, 14, 14, 14]
        for col, width in enumerate(widths, 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        sheet.freeze_panes = "A3"
        workbook.save(path)
        try:
            os.startfile(path)
        except OSError as exc:
            messagebox.showinfo("엑셀 미리보기", f"엑셀 파일을 만들었지만 자동으로 열지 못했습니다.\n\n{path}\n\n{exc}")


if __name__ == "__main__":
    app = SearchApp()
    app.mainloop()
