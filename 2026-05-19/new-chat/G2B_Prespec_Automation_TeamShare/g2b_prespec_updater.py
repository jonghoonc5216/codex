#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
G2B preliminary specification updater.

Reads Korea Public Procurement Service pre-spec data, merges missed dates since
the last successful run, writes a focused Excel workbook, and optionally opens it.
"""

from __future__ import annotations

import argparse
import hashlib
import html
import json
import os
import re
import sys
import time
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import quote, quote_plus
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


APP_DIR = Path(__file__).resolve().parent
CONFIG_PATH = APP_DIR / "config.json"
STATE_PATH = APP_DIR / "state.json"
RECORDS_PATH = APP_DIR / "records.json"
LOG_PATH = APP_DIR / "g2b_prespec_update.log"

API_URL = (
    "http://apis.data.go.kr/1230000/ao/"
    "HrcspSsstndrdInfoService/getPublicPrcureThngInfoServc"
)
G2B_DETAIL_URL = "https://www.g2b.go.kr/"

DEFAULT_CONFIG = {
    "api_key": "",
    "search_keyword": "공원",
    "search_keywords": [
        "공원",
        "파크",
        "조경",
        "관광",
        "유원지",
        "체육",
        "탐방로",
        "수목원",
        "정원",
        "숲",
        "휴양림",
        "치유",
        "야영장",
        "레포츠",
        "산림",
        "체험",
        "생태",
        "경관",
        "도시재생",
        "바람길",
        "둘레길",
    ],
    "exclude_keywords": ["위탁", "건설폐기물", "임차", "감리", "건설사업관리"],
    "business_categories": ["일반용역", "기술용역"],
    "output_excel": "조달청_사전규격_자동수집.xlsx",
    "seed_excel": str(Path.home() / "Downloads" / "조달청 사전규격.xlsx"),
    "overlap_days": 1,
    "chunk_days": 15,
    "num_of_rows": 999,
    "display_mode": "since_last_success",
    "open_excel_after_update": True,
}

HEADERS = [
    "신규",
    "No",
    "사전규격등록번호",
    "업무구분",
    "사업명",
    "수요기관",
    "담당자",
    "진행일자",
    "의견등록 마감일시",
    "진행상태",
    "예산금액",
]

WIDTHS = {
    "A": 8,
    "B": 6,
    "C": 18,
    "D": 12,
    "E": 52,
    "F": 24,
    "G": 10,
    "H": 13,
    "I": 19,
    "J": 10,
    "K": 16,
}


def log(message: str) -> None:
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    text = f"[{stamp}] {message}"
    print(text)
    with LOG_PATH.open("a", encoding="utf-8") as fh:
        fh.write(text + "\n")


def read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    with path.open("r", encoding="utf-8-sig") as fh:
        return json.load(fh)


def write_json(path: Path, data: Any) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)
    tmp.replace(path)


def expand_path(value: str) -> Path:
    path = Path(os.path.expandvars(value)).expanduser()
    if not path.is_absolute():
        return APP_DIR / path
    return path


def ensure_config() -> dict[str, Any]:
    if not CONFIG_PATH.exists():
        write_json(CONFIG_PATH, DEFAULT_CONFIG)
        log(f"Created config: {CONFIG_PATH}")
    config = DEFAULT_CONFIG | read_json(CONFIG_PATH, {})
    config["api_key"] = (
        config.get("api_key")
        or os.environ.get("NARA_PRESPEC_API_KEY")
        or os.environ.get("NARA_API_KEY")
        or ""
    )
    return config


def get_search_keywords(config: dict[str, Any]) -> list[str]:
    raw_keywords = config.get("search_keywords")
    keywords: list[str] = []
    if isinstance(raw_keywords, list):
        keywords.extend(normalize_text(value) for value in raw_keywords)
    elif isinstance(raw_keywords, str):
        keywords.extend(normalize_text(value) for value in re.split(r"[,\\s]+", raw_keywords))
    legacy_keyword = normalize_text(config.get("search_keyword", ""))
    if legacy_keyword:
        keywords.append(legacy_keyword)
    deduped: list[str] = []
    for keyword in keywords:
        if keyword and keyword not in deduped:
            deduped.append(keyword)
    return deduped


def get_exclude_keywords(config: dict[str, Any]) -> list[str]:
    raw_keywords = config.get("exclude_keywords", [])
    keywords: list[str] = []
    if isinstance(raw_keywords, list):
        keywords.extend(normalize_text(value) for value in raw_keywords)
    elif isinstance(raw_keywords, str):
        keywords.extend(normalize_text(value) for value in re.split(r"[,\\s]+", raw_keywords))
    return [keyword for keyword in keywords if keyword]


def record_matches_keywords(record: dict[str, Any], config: dict[str, Any]) -> bool:
    title = normalize_text(record.get("사업명"))
    includes = get_search_keywords(config)
    excludes = get_exclude_keywords(config)
    if includes and not any(keyword in title for keyword in includes):
        return False
    if excludes and any(keyword in title for keyword in excludes):
        return False
    return True


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = html.unescape(str(value)).replace("\u200b", "").strip()
    return re.sub(r"\s+", " ", text)


def pick(item: dict[str, Any], *names: str, default: str = "") -> str:
    for name in names:
        value = item.get(name)
        if value not in (None, ""):
            return normalize_text(value)
    return default


def money_text(value: Any) -> str:
    raw = normalize_text(value).replace(",", "")
    if not raw:
        return ""
    try:
        return f"{int(float(raw)):,}"
    except ValueError:
        return normalize_text(value)


def parse_date(value: Any) -> date | None:
    text = normalize_text(value)
    if not text:
        return None
    digits = re.sub(r"\D", "", text)
    if len(digits) >= 8:
        try:
            return datetime.strptime(digits[:8], "%Y%m%d").date()
        except ValueError:
            return None
    return None


def date_text(value: Any) -> str:
    parsed = parse_date(value)
    return parsed.strftime("%Y/%m/%d") if parsed else normalize_text(value)


def datetime_text(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    digits = re.sub(r"\D", "", text)
    for fmt, size, out_fmt in (
        ("%Y%m%d%H%M", 12, "%Y/%m/%d %H:%M"),
        ("%Y%m%d", 8, "%Y/%m/%d"),
    ):
        if len(digits) >= size:
            try:
                return datetime.strptime(digits[:size], fmt).strftime(out_fmt)
            except ValueError:
                pass
    return text


def parse_datetime(value: Any) -> datetime | None:
    text = normalize_text(value)
    digits = re.sub(r"\D", "", text)
    for fmt, size in (("%Y%m%d%H%M", 12), ("%Y%m%d", 8)):
        if len(digits) >= size:
            try:
                return datetime.strptime(digits[:size], fmt)
            except ValueError:
                pass
    return None


def record_key(record: dict[str, Any]) -> str:
    spec_no = normalize_text(record.get("사전규격등록번호"))
    if spec_no:
        return f"spec:{spec_no}"
    return "seed:" + natural_key(record)


def loose_key(record: dict[str, Any]) -> str:
    parts = [
        normalize_text(record.get("사업명")),
        normalize_text(record.get("수요기관")),
        normalize_text(record.get("공고기관")),
    ]
    raw = "|".join(parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def natural_key(record: dict[str, Any]) -> str:
    parts = [
        normalize_text(record.get("사업명")),
        normalize_text(record.get("수요기관")),
        normalize_text(record.get("공고기관")),
        normalize_text(record.get("진행일자")),
    ]
    raw = "|".join(parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def detail_url(record: dict[str, Any]) -> str:
    spec_no = normalize_text(record.get("사전규격등록번호"))
    if not spec_no:
        return ""
    business_name = normalize_text(record.get("업무구분"))
    business_code = {
        "물품": "01",
        "일반용역": "02",
        "기술용역": "03",
        "공사": "04",
        "외자": "05",
        "리스": "06",
    }.get(business_name, "")
    params = {
        "g2bPrespecRegNo": spec_no,
        "bfSpecRegNo": spec_no,
        "bfSpecRgstNo": spec_no,
        "bsnsDivNm": business_name,
    }
    if business_code:
        params["prcmBsneSeCd"] = business_code
    return G2B_DETAIL_URL + "?" + build_query(params)


def detect_business_category(item: dict[str, Any]) -> str:
    text = pick(
        item,
        "bsnsDivNm",
        "bizDivNm",
        "businessDivNm",
        "bfSpecBsnsDivNm",
        "bsnsSeNm",
        "업무구분",
    )
    if text:
        return text
    service_name = pick(item, "prdctClsfcNoNm", "품명", "사업명")
    if "설계" in service_name or "감리" in service_name or "건설사업관리" in service_name:
        return "기술용역"
    return "일반용역"


def normalize_api_item(item: dict[str, Any]) -> dict[str, Any]:
    category = detect_business_category(item)
    deadline_raw = pick(
        item,
        "opninRgstClseDt",
        "opninRgstClseDttm",
        "opninRgstEndDt",
        "opninRgstClseDate",
        "opinionRgstClseDt",
        "opinionRegistClseDt",
        "의견등록 마감일시",
    )
    deadline_dt = parse_datetime(deadline_raw)
    progress_raw = pick(
        item,
        "bfSpecRgstDt",
        "rgstDt",
        "rcptDt",
        "rlsDt",
        "registDt",
        "writngDt",
        "inqryDt",
        default=deadline_raw,
    )
    status = pick(item, "bfSpecPrgsSttsNm", "prgsSttsNm", "진행상태")
    if not status:
        status = "게시중" if (deadline_dt is None or deadline_dt >= datetime.now()) else "마감"
    business_yn = pick(item, "dmstcDlvryYn", "indstrytyNm", "업무여부")
    if not business_yn:
        business_yn = "내자" if category == "일반용역" else "-"
    return {
        "사전규격등록번호": pick(
            item,
            "bfSpecRgstNo",
            "bfSpecRegNo",
            "preSpecRgstNo",
            "preSpecRegNo",
            "preSpecNo",
            "사전규격등록번호",
        ),
        "업무구분": category,
        "업무여부": business_yn,
        "사업명": pick(item, "prdctClsfcNoNm", "specNm", "bidNtceNm", "사업명"),
        "수요기관": pick(item, "dminsttNm", "rlDminsttNm", "demandInsttNm", "수요기관"),
        "공고기관": pick(item, "orderInsttNm", "ntceInsttNm", "공고기관"),
        "담당자": pick(item, "ofclNm", "chargerNm", "담당자", default="-"),
        "진행일자": date_text(progress_raw),
        "의견등록 마감일시": datetime_text(deadline_raw),
        "진행상태": status,
        "참조여부": pick(item, "rfrncYn", "refYn", "참조여부"),
        "업체등록수": pick(item, "entrpsRgstCnt", "corpRgstCnt", "업체등록수", default="0"),
        "예산금액": money_text(pick(item, "asignBdgtAmt", "bdgtAmt", "budgetAmt", "예산금액")),
        "_raw": item,
        "_source": "api",
    }


def category_allowed(record: dict[str, Any], allowed: list[str]) -> bool:
    if not allowed:
        return True
    category = normalize_text(record.get("업무구분"))
    return any(name in category for name in allowed)


def build_query(params: dict[str, Any]) -> str:
    parts = []
    for key, value in params.items():
        if value is None or value == "":
            continue
        if key.lower() == "servicekey":
            encoded = quote(str(value), safe="%")
        else:
            encoded = quote_plus(str(value))
        parts.append(f"{quote_plus(key)}={encoded}")
    return "&".join(parts)


def request_json(params: dict[str, Any], retries: int = 2) -> dict[str, Any]:
    url = API_URL + "?" + build_query(params)
    headers = {"User-Agent": "Mozilla/5.0 g2b-prespec-updater"}
    last_error: Exception | None = None
    for attempt in range(retries + 1):
        try:
            with urlopen(Request(url, headers=headers), timeout=35) as resp:
                raw = resp.read().decode("utf-8-sig")
            data = json.loads(raw)
            return data
        except Exception as exc:
            last_error = exc
            if attempt < retries:
                time.sleep(2 + attempt)
    raise RuntimeError(f"API request failed: {last_error}")


def items_from_response(data: dict[str, Any]) -> tuple[list[dict[str, Any]], int]:
    response = data.get("response", {})
    header = response.get("header", {})
    code = str(header.get("resultCode", ""))
    if code and code != "00":
        msg = header.get("resultMsg", "")
        raise RuntimeError(f"API error {code}: {msg}")
    body = response.get("body", {})
    total = int(body.get("totalCount") or 0)
    items = body.get("items") or []
    if isinstance(items, dict):
        items = items.get("item") or []
    if isinstance(items, dict):
        items = [items]
    if isinstance(items, str):
        items = []
    return [x for x in items if isinstance(x, dict)], total


def date_chunks(start: date, end: date, chunk_days: int) -> list[tuple[date, date]]:
    chunks = []
    cur = start
    while cur <= end:
        chunk_end = min(cur + timedelta(days=max(chunk_days, 1) - 1), end)
        chunks.append((cur, chunk_end))
        cur = chunk_end + timedelta(days=1)
    return chunks


def fetch_api_records(config: dict[str, Any], start: date, end: date) -> list[dict[str, Any]]:
    api_key = config.get("api_key", "")
    if not api_key:
        raise RuntimeError("config.json에 공공데이터포털 API 키가 없습니다.")
    keywords = get_search_keywords(config)
    # This API silently falls back to 10 rows when numOfRows is 1000 or higher.
    rows_per_page = min(max(int(config.get("num_of_rows", 999)), 1), 999)
    all_items: list[dict[str, Any]] = []
    for chunk_start, chunk_end in date_chunks(start, end, int(config.get("chunk_days", 15))):
        page = 1
        while True:
            params = {
                "serviceKey": api_key,
                "type": "json",
                "inqryDiv": "1",
                "inqryBgnDt": chunk_start.strftime("%Y%m%d0000"),
                "inqryEndDt": chunk_end.strftime("%Y%m%d2359"),
                "pageNo": page,
                "numOfRows": rows_per_page,
            }
            data = request_json(params)
            items, total = items_from_response(data)
            all_items.extend(items)
            if not items or page * rows_per_page >= total:
                break
            page += 1
    allowed = list(config.get("business_categories") or [])
    records = [normalize_api_item(item) for item in all_items]
    if keywords or get_exclude_keywords(config):
        records = [record for record in records if record_matches_keywords(record, config)]
    return [record for record in records if category_allowed(record, allowed)]


def find_header_row(ws: Any) -> int | None:
    for row in range(1, ws.max_row + 1):
        values = [normalize_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if "No" in values and "사업명" in values:
            return row
    return None


def import_seed_records(seed_path: Path) -> list[dict[str, Any]]:
    if not seed_path.exists():
        return []
    wb = load_workbook(seed_path, data_only=True)
    records: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        header_row = find_header_row(ws)
        if not header_row:
            continue
        headers = [normalize_text(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
        for row in range(header_row + 1, ws.max_row + 1):
            values = {
                headers[col - 1]: normalize_text(ws.cell(row, col).value)
                for col in range(1, ws.max_column + 1)
                if headers[col - 1]
            }
            if not values.get("사업명"):
                continue
            category = values.get("업무구분", "")
            record = {
                "사전규격등록번호": values.get("사전규격등록번호", ""),
                "업무구분": category,
                "업무여부": values.get("업무여부", "내자" if category == "일반용역" else "-"),
                "사업명": values.get("사업명", ""),
                "수요기관": values.get("수요기관", ""),
                "공고기관": values.get("공고기관", ""),
                "담당자": values.get("담당자", "-"),
                "진행일자": date_text(values.get("진행일자", "")),
                "의견등록 마감일시": values.get("의견등록 마감일시", ""),
                "진행상태": values.get("진행상태", ""),
                "참조여부": values.get("참조여부", ""),
                "업체등록수": values.get("업체등록수", "0"),
                "예산금액": money_text(values.get("예산금액", "")),
                "_source": "seed",
            }
            record["_key"] = record_key(record)
            record["_natural_key"] = natural_key(record)
            record["_first_seen"] = ""
            records.append(record)
    return records


def merge_records(
    current: list[dict[str, Any]],
    incoming: list[dict[str, Any]],
    mark_new: bool,
) -> tuple[list[dict[str, Any]], int]:
    by_key: dict[str, dict[str, Any]] = {}
    natural_to_key: dict[str, str] = {}
    loose_to_key: dict[str, str] = {}
    for record in current:
        key = record.get("_key") or record_key(record)
        natural = record.get("_natural_key") or natural_key(record)
        loose = record.get("_loose_key") or loose_key(record)
        record["_key"] = key
        record["_natural_key"] = natural
        record["_loose_key"] = loose
        by_key[key] = record
        natural_to_key[natural] = key
        loose_to_key[loose] = key

    today_text = date.today().isoformat()
    new_count = 0
    for record in incoming:
        key = record_key(record)
        natural = natural_key(record)
        loose = loose_key(record)
        existing_key = key if key in by_key else natural_to_key.get(natural) or loose_to_key.get(loose)
        is_new = existing_key is None
        merged = by_key.get(existing_key, {}) if existing_key else {}
        merged.update(record)
        merged["_key"] = key
        merged["_natural_key"] = natural
        merged["_loose_key"] = loose
        if is_new and mark_new:
            merged["_first_seen"] = today_text
            new_count += 1
        elif not merged.get("_first_seen"):
            merged["_first_seen"] = ""
        if existing_key and existing_key != key:
            by_key.pop(existing_key, None)
        by_key[key] = merged
        natural_to_key[natural] = key
        loose_to_key[loose] = key
    return list(by_key.values()), new_count


def sort_key(record: dict[str, Any]) -> tuple[int, int, str]:
    is_new = 0 if record.get("_first_seen") == date.today().isoformat() else 1
    progress = parse_date(record.get("진행일자")) or date.min
    return (is_new, -progress.toordinal(), normalize_text(record.get("사업명")))


def parse_state_date(value: Any) -> date | None:
    text = normalize_text(value)
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        return None


def choose_display_period(state: dict[str, Any], today: date, config: dict[str, Any]) -> tuple[date, date]:
    mode = normalize_text(config.get("display_mode", "since_last_success")).lower()
    if mode in {"week", "weekly", "this_week"}:
        return today - timedelta(days=today.weekday()), today

    last = parse_state_date(state.get("last_success_date"))
    if last == today:
        stored_display_start = parse_state_date(state.get("last_display_start_date"))
        stored_highlight_start = parse_state_date(state.get("last_highlight_start_date"))
        if stored_display_start and stored_display_start <= today:
            if stored_highlight_start and stored_display_start >= stored_highlight_start:
                return stored_highlight_start - timedelta(days=1), today
            return stored_display_start, today
        if stored_highlight_start:
            return stored_highlight_start - timedelta(days=1), today
        return today - timedelta(days=1), today

    carry_start = parse_state_date(state.get("next_display_start_date"))
    if carry_start and carry_start <= today:
        return carry_start, today

    if last and last <= today:
        return last, today
    return today - timedelta(days=1), today


def choose_highlight_start(state: dict[str, Any], today: date) -> date:
    last = parse_state_date(state.get("last_success_date"))
    if last and last < today:
        return last + timedelta(days=1)
    if last == today:
        stored = parse_state_date(state.get("last_highlight_start_date"))
        if stored and stored <= today:
            return stored
    return today


def sheet_title_for_period(start: date, end: date) -> str:
    if start == end:
        return f"{end:%Y.%m.%d.}"
    if start.year == end.year:
        return f"{start:%Y.%m.%d}-{end:%m.%d}"
    return f"{start:%Y.%m.%d}-{end:%Y.%m.%d}"


def is_highlight_record(record: dict[str, Any], highlight_start: date, period_end: date) -> bool:
    progress = parse_date(record.get("진행일자"))
    return bool(progress and highlight_start <= progress <= period_end)


def period_sort_key(record: dict[str, Any], highlight_start: date, period_end: date) -> tuple[int, int, str]:
    is_highlight = 0 if is_highlight_record(record, highlight_start, period_end) else 1
    progress = parse_date(record.get("진행일자")) or date.min
    return (is_highlight, -progress.toordinal(), normalize_text(record.get("사업명")))


def records_in_period(
    records: list[dict[str, Any]],
    start: date,
    end: date,
    config: dict[str, Any],
) -> list[dict[str, Any]]:
    filtered: list[dict[str, Any]] = []
    for record in records:
        progress = parse_date(record.get("진행일자"))
        if progress and start <= progress <= end and record_matches_keywords(record, config):
            filtered.append(record)
    return filtered


def write_sheet(
    wb: Workbook,
    title: str,
    records: list[dict[str, Any]],
    config: dict[str, Any],
    today: date,
    period_start: date,
    period_end: date,
    highlight_start: date,
) -> None:
    ws = wb.create_sheet(title)
    keyword = normalize_text(config.get("search_keyword"))
    keywords = get_search_keywords(config)
    excludes = get_exclude_keywords(config)
    keyword_text = ", ".join(keywords) if keywords else keyword
    exclude_text = ", ".join(excludes)
    categories = " ".join(config.get("business_categories") or [])

    info = [
        ("발주목록_사전규격공개", ""),
        ("검색유형", "사전규격공개"),
        ("사업명", keyword_text),
        ("제외단어", exclude_text),
        ("진행일자", f"{period_start:%Y%m%d} - {period_end:%Y%m%d}"),
        ("업무구분", categories),
    ]
    title_fill = PatternFill("solid", fgColor="EAF2FF")
    label_fill = PatternFill("solid", fgColor="F2F5F9")
    header_fill = PatternFill("solid", fgColor="DDEBFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row, (label, value) in enumerate(info, 1):
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 1).fill = title_fill if row == 1 else label_fill
        ws.cell(row, 1).alignment = Alignment(horizontal="left")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=len(HEADERS))

    header_row = 8
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(header_row, col, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    red_font = Font(color="C00000", bold=True)
    normal_font = Font(color="000000")
    hyperlink_font = Font(color="0563C1", underline="single")
    new_hyperlink_font = Font(color="C00000", bold=True, underline="single")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    sorted_records = sorted(records, key=lambda record: period_sort_key(record, highlight_start, period_end))
    for idx, record in enumerate(sorted_records, 1):
        row = header_row + idx
        is_new = is_highlight_record(record, highlight_start, period_end)
        values = [
            "신규" if is_new else "",
            idx,
            record.get("사전규격등록번호", ""),
            record.get("업무구분", ""),
            record.get("사업명", ""),
            record.get("수요기관", ""),
            record.get("담당자", ""),
            record.get("진행일자", ""),
            record.get("의견등록 마감일시", ""),
            record.get("진행상태", ""),
            record.get("예산금액", ""),
        ]
        link = detail_url(record)
        for col, value in enumerate(values, 1):
            cell = ws.cell(row, col, value)
            cell.border = border
            cell.font = copy(red_font if is_new else normal_font)
            if col == 5 and link:
                cell.hyperlink = link
                cell.font = copy(new_hyperlink_font if is_new else hyperlink_font)
            if col in (1, 2, 4, 7, 8, 10):
                cell.alignment = center
            elif col == 11:
                cell.alignment = right
            else:
                cell.alignment = left

    for letter, width in WIDTHS.items():
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A9"
    ws.auto_filter.ref = f"A8:{get_column_letter(len(HEADERS))}{max(8, ws.max_row)}"
    ws.sheet_view.showGridLines = False


def write_workbook(
    records: list[dict[str, Any]],
    output_path: Path,
    config: dict[str, Any],
    period_start: date,
    period_end: date,
    highlight_start: date,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    today = date.today()
    title = sheet_title_for_period(period_start, period_end)
    write_sheet(
        wb,
        title,
        records_in_period(records, period_start, period_end, config),
        config,
        today,
        period_start,
        period_end,
        highlight_start,
    )
    wb.save(output_path)


def choose_start_date(state: dict[str, Any], overlap_days: int) -> date:
    last = parse_state_date(state.get("last_success_date"))
    if last:
        if overlap_days <= 0:
            return last + timedelta(days=1)
        return last - timedelta(days=max(overlap_days - 1, 0))
    today = date.today()
    return today


def open_excel(path: Path) -> None:
    if sys.platform.startswith("win"):
        os.startfile(str(path))  # type: ignore[attr-defined]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--no-open", action="store_true", help="Do not open Excel after update")
    args = parser.parse_args()

    config = ensure_config()
    output_path = expand_path(config["output_excel"])
    seed_path = expand_path(config["seed_excel"])
    state = read_json(STATE_PATH, {})
    records = read_json(RECORDS_PATH, [])

    if not records:
        seed_records = import_seed_records(seed_path)
        if seed_records:
            records, _ = merge_records([], seed_records, mark_new=False)
            log(f"Imported seed Excel rows: {len(seed_records)}")

    today = date.today()
    display_start, display_end = choose_display_period(state, today, config)
    highlight_start = choose_highlight_start(state, today)
    fetched = []
    success = False
    try:
        start = choose_start_date(state, int(config.get("overlap_days", 1)))
        if start > today:
            start = today
        log(f"Collecting API records: {start} ~ {today}")
        fetched = fetch_api_records(config, start, today)
        mark_new = bool(records or state.get("last_success_date"))
        records, new_count = merge_records(records, fetched, mark_new=mark_new)
        state["last_success_date"] = today.isoformat()
        state["last_success_at"] = datetime.now().isoformat(timespec="seconds")
        state["last_fetch_count"] = len(fetched)
        state["last_new_count"] = new_count
        state["last_display_start_date"] = display_start.isoformat()
        state["last_highlight_start_date"] = highlight_start.isoformat()
        state["next_display_start_date"] = highlight_start.isoformat()
        success = True
        log(f"API records fetched: {len(fetched)}, new: {new_count}")
    except Exception as exc:
        state["last_error_at"] = datetime.now().isoformat(timespec="seconds")
        state["last_error"] = str(exc)
        log(f"API update skipped/failed: {exc}")

    write_json(RECORDS_PATH, records)
    write_json(STATE_PATH, state)
    write_workbook(records, output_path, config, display_start, display_end, highlight_start)
    log(f"Excel display period: {display_start} ~ {display_end}")
    log(f"Excel highlight period: {highlight_start} ~ {display_end}")
    log(f"Excel written: {output_path}")

    if config.get("open_excel_after_update", True) and not args.no_open:
        open_excel(output_path)
    return 0 if success or records else 1


if __name__ == "__main__":
    raise SystemExit(main())
