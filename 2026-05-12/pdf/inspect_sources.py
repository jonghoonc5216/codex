from pathlib import Path

import openpyxl
import pdfplumber


PDF_PATH = Path(r"C:\Users\saman\Desktop\Ai 실습\_새 폴더\실시계획인가.pdf")
XLSX_PATH = Path(r"C:\Users\saman\Desktop\Ai 실습\_새 폴더\260127 실시계획인가 금액산정.xlsx")


def show_pdf_tables() -> None:
    print("## PDF tables")
    with pdfplumber.open(PDF_PATH) as pdf:
        print("pages:", len(pdf.pages))
        for page_index, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            print(f"page {page_index}: {len(tables)} table(s)")
            for table_index, table in enumerate(tables, start=1):
                col_count = max((len(row) for row in table), default=0)
                print(f"  table {table_index}: {len(table)} row(s), {col_count} col(s)")
                for row in table[:20]:
                    print("    " + " | ".join((cell or "").replace("\n", "/") for cell in row))


def show_workbook() -> None:
    print("## Workbook")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    print("sheets:", wb.sheetnames)
    for ws in wb.worksheets:
        print(f"\n--- {ws.title} ({ws.max_row} x {ws.max_column}) ---")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            values = []
            for cell in row:
                value = cell.value
                values.append("" if value is None else str(value))
            print("\t".join(values))


if __name__ == "__main__":
    show_pdf_tables()
    show_workbook()
