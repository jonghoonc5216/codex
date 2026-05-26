from pathlib import Path

import openpyxl

from pdf_formula_converter import ExtractResult, fill_workload_sheet, parse_work_items


table = [
    ["기본업무", "단위", "기술사", "특급", "고급", "중급", "초급", "환산"],
    ["1. 조사 및 분석", "별도 산정", "", "", "", "", "", ""],
    ["2. 기본구상", "", "", "", "", "", "", ""],
    ["2.1 계획의 배경", "5만인", "4.9", "6.9", "16.2", "23.5", "26.0", "①"],
    ["2.2 시ㆍ군의 장기발전구상", "5만인", "7.6", "11.1", "29.3", "43.5", "26.0", "①"],
    ["3. 부문별계획", "", "", "", "", "", "", ""],
    ["3.4 경관계획", "별도 산정", "", "", "", "", "", ""],
    ["3.5 단계별 집행계획", "별도 산정", "", "", "", "", "", ""],
]

items = parse_work_items(table)
for item in items:
    print(item.label, "display", item.display_only, item.values)
    for child in item.sub_items:
        print("  ", child.label, "display", child.display_only, child.values)

wb = openpyxl.load_workbook(Path("template.xlsx"))
ws = wb["실시계획인가 소요작업량"]
result = ExtractResult(Path("test.pdf"), 1, [], [], [], items, [])
summary = fill_workload_sheet(ws, result, 1445, factor_formula_override="=ROUND((D2/1)^(0.6),3)", ratio=1)
print("summary", summary)
for row in range(7, summary + 1):
    print(row, ws.cell(row, 2).value, ws.cell(row, 3).value, ws.cell(row, 9).value, ws.cell(row, 15).value)
