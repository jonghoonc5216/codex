from __future__ import annotations

import argparse
import copy
import datetime as dt
import math
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

import openpyxl
import pdfplumber
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter


DEFAULT_TEMPLATE = Path(r"C:\Users\saman\Desktop\Ai 실습\_새 폴더\260127 실시계획인가 금액산정.xlsx")
DEFAULT_PDF = Path(r"C:\Users\saman\Desktop\Ai 실습\_새 폴더\실시계획인가.pdf")
DEFAULT_OUTPUT = Path(r"C:\Users\saman\Desktop\Ai 실습\_새 폴더\PDF표_산식변환_결과.xlsx")

AMOUNT_TEMPLATE_SHEET = "실시계획인가"
WORKLOAD_TEMPLATE_SHEET = "실시계획인가 소요작업량"

INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\']")


@dataclass
class WorkSubItem:
    label: str
    unit: str
    values: list[float]
    factor_symbol: str
    display_only: bool = False


@dataclass
class WorkItem:
    label: str
    unit: str
    values: list[float]
    factor_symbol: str
    sub_items: list[WorkSubItem] = field(default_factory=list)
    display_only: bool = False


@dataclass
class ExtractResult:
    pdf_path: Path
    page_count: int
    raw_tables: list[list[list[str]]]
    header_lines: list[str]
    note_lines: list[str]
    items: list[WorkItem]
    warnings: list[str]
    raw_sheet_refs: dict[str, list[str]] = field(default_factory=dict)
    amount_entries: list[dict[str, object]] = field(default_factory=list)


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip()


def parse_number(value: object) -> float | None:
    text = normalize_text(value).replace(",", "").replace("−", "-")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group(0))


def clean_task_label(label: str) -> str:
    label = normalize_text(label)
    label = re.sub(r"^\((\d+)\)\s*", r"\1. ", label)
    label = re.sub(r"^(\d+)\)\s*", r"\1. ", label)
    return label


def workload_label(item: WorkItem) -> str:
    if not item.sub_items:
        return item.label
    labels = [item.label]
    labels.extend(child.label for child in item.sub_items if child.label and child.label not in labels)
    return "\n".join(labels)


def parse_hierarchy(label: str) -> tuple[str | None, int]:
    match = re.match(r"^\s*\(?(\d+(?:\.\d+)*)\)?\s*[.)]?\s*", normalize_text(label))
    if not match:
        return None, 0
    parts = match.group(1).split(".")
    return parts[0], len(parts)


def title_without_leading_number(title: str) -> str:
    text = normalize_text(title)
    stripped = re.sub(r"^\s*(?:\(?\d+\)?\s*(?:번)?\s*[._)\-\s]*)+", "", text).strip()
    return stripped or text


def coerce_cell_text(cell: object) -> str:
    return "" if cell is None else str(cell)


def table_has_expected_shape(table: list[list[object]]) -> bool:
    if len(table) < 3:
        return False
    max_cols = max((len(row) for row in table), default=0)
    if max_cols < 7:
        return False
    joined = " ".join(normalize_text(cell) for row in table[:3] for cell in row)
    return "기본업무" in joined or "기준인원수" in joined or "기술자" in joined


def split_pdf_context(text: str) -> tuple[list[str], list[str]]:
    lines = [normalize_text(line) for line in text.splitlines()]
    lines = [line for line in lines if line]

    table_start = None
    for index, line in enumerate(lines):
        if "기본업무" in line or "기준인원수" in line:
            table_start = index
            break
    header_lines = lines[:table_start] if table_start is not None else []

    note_start = None
    for index, line in enumerate(lines):
        if re.match(r"^주\s*\d+\)", line) or line.startswith("주 "):
            note_start = index
            break
    note_lines: list[str] = []
    if note_start is not None:
        for line in lines[note_start:]:
            if re.fullmatch(r"-\s*\d+\s*-", line):
                continue
            note_lines.append(line)

    return header_lines, note_lines


def parse_work_items(table: list[list[object]]) -> list[WorkItem]:
    entries: list[dict[str, object]] = []
    parent_labels: dict[str, str] = {}
    parent_order: list[str] = []
    child_tops: set[str] = set()

    for row in table:
        padded = list(row) + [""] * max(0, 8 - len(row))
        raw_label = normalize_text(padded[0])
        label = clean_task_label(raw_label)
        top_id, level = parse_hierarchy(raw_label)
        if top_id and level == 1 and label:
            parent_labels[top_id] = label
            if top_id not in parent_order:
                parent_order.append(top_id)

        numbers = [parse_number(padded[index]) for index in range(2, 7)]
        has_values = sum(number is not None for number in numbers) >= 3
        unit = normalize_text(padded[1])
        should_keep_display_row = bool(label and top_id and not has_values and (unit or level >= 1))
        if not label or (not has_values and not should_keep_display_row):
            continue
        output_label = label
        values = [0.0 if number is None else number for number in numbers]
        if top_id and top_id not in parent_order:
            parent_order.append(top_id)
        if top_id and level > 1:
            child_tops.add(top_id)
        entries.append(
            {
                "label": output_label,
                "raw_label": raw_label,
                "unit": unit,
                "values": values,
                "factor_symbol": normalize_text(padded[7]),
                "top_id": top_id,
                "level": level,
                "has_values": has_values,
            }
        )

    items: list[WorkItem] = []
    used_top_ids: set[str] = set()
    for top_id in parent_order:
        group_entries = [entry for entry in entries if entry["top_id"] == top_id]
        if not group_entries:
            continue
        if top_id not in child_tops:
            for entry in group_entries:
                items.append(
                    WorkItem(
                        label=str(entry["label"]),
                        unit=str(entry["unit"]),
                        values=[float(value) for value in entry["values"]],
                        factor_symbol=str(entry["factor_symbol"]),
                        display_only=not bool(entry["has_values"]),
                    )
                )
            used_top_ids.add(top_id)
            continue

        first = next((entry for entry in group_entries if entry["has_values"]), group_entries[0])
        sub_items = [
            WorkSubItem(
                label=str(entry["raw_label"]),
                unit=str(entry["unit"]),
                values=[float(value) for value in entry["values"]],
                factor_symbol=str(entry["factor_symbol"]),
                display_only=not bool(entry["has_values"]),
            )
            for entry in group_entries
            if int(entry["level"]) > 1 and normalize_text(entry["raw_label"])
        ]
        items.append(
            WorkItem(
                label=parent_labels.get(top_id, str(first["label"])),
                unit=str(first["unit"]),
                values=[0.0] * 5,
                factor_symbol=str(first["factor_symbol"]),
                sub_items=sub_items,
                display_only=True,
            )
        )
        used_top_ids.add(top_id)

    for entry in entries:
        top_id = entry["top_id"]
        if top_id in used_top_ids:
            continue
        items.append(
            WorkItem(
                label=str(entry["label"]),
                unit=str(entry["unit"]),
                values=[float(value) for value in entry["values"]],
                factor_symbol=str(entry["factor_symbol"]),
                display_only=not bool(entry["has_values"]),
            )
        )
    return items


def extract_pdf(pdf_path: Path) -> ExtractResult:
    raw_tables: list[list[list[str]]] = []
    header_lines: list[str] = []
    note_lines: list[str] = []
    items: list[WorkItem] = []
    warnings: list[str] = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_index, page in enumerate(pdf.pages, start=1):
            page_header_lines, page_note_lines = split_pdf_context(page.extract_text() or "")
            header_lines.extend(line for line in page_header_lines if line not in header_lines)
            note_lines.extend(line for line in page_note_lines if line not in note_lines)
            tables = page.extract_tables()
            if not tables:
                warnings.append(f"{page_index}쪽에서 표를 찾지 못했습니다.")
                continue
            matched_on_page = 0
            for table in tables:
                text_table = [[coerce_cell_text(cell) for cell in row] for row in table]
                raw_tables.append(text_table)
                if not table_has_expected_shape(table):
                    continue
                parsed = parse_work_items(table)
                if parsed:
                    matched_on_page += 1
                    items.extend(parsed)
            if matched_on_page == 0:
                warnings.append(f"{page_index}쪽 표는 예상 양식과 달라 수식 표로 반영하지 못했습니다.")

        page_count = len(pdf.pages)

    if not items:
        raise ValueError(
            f"{pdf_path.name}에서 '기본업무/기준인원수' 형태의 표를 찾지 못했습니다. "
            "스캔 이미지 PDF라면 OCR 단계가 추가로 필요합니다."
        )

    return ExtractResult(
        pdf_path=pdf_path,
        page_count=page_count,
        raw_tables=raw_tables,
        header_lines=header_lines,
        note_lines=note_lines,
        items=items,
        warnings=warnings,
    )


def sanitize_sheet_base(name: str) -> str:
    name = INVALID_SHEET_CHARS.sub("_", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name or "PDF"


def unique_sheet_name(workbook: openpyxl.Workbook, desired: str) -> str:
    desired = desired[:31]
    if desired not in workbook.sheetnames:
        return desired
    for index in range(2, 1000):
        suffix = f"_{index}"
        candidate = f"{desired[:31 - len(suffix)]}{suffix}"
        if candidate not in workbook.sheetnames:
            return candidate
    raise ValueError("사용 가능한 엑셀 시트 이름을 만들지 못했습니다.")


def make_sheet_name(workbook: openpyxl.Workbook, base: str, suffix: str) -> str:
    base_limit = 31 - len(suffix)
    return unique_sheet_name(workbook, f"{base[:base_limit]}{suffix}")


def quote_sheet_ref(sheet_name: str) -> str:
    return f"'{sheet_name.replace(chr(39), chr(39) * 2)}'!"


def find_workload_block(sheet) -> tuple[int, int, int]:
    summary_row = None
    for row in range(1, sheet.max_row + 1):
        label = normalize_text(sheet.cell(row=row, column=2).value).replace(" ", "")
        if label == "소계":
            summary_row = row
            break
    if summary_row is None:
        return 7, 11, 12

    data_start = None
    for row in range(1, summary_row):
        label = normalize_text(sheet.cell(row=row, column=2).value)
        if not label or not re.search(r"[가-힣A-Za-z]", label):
            continue
        values = [parse_number(sheet.cell(row=row, column=column).value) for column in range(3, 8)]
        if sum(value is not None for value in values) >= 3:
            data_start = row
            break
    if data_start is None:
        data_start = max(1, summary_row - 5)
    return data_start, summary_row - 1, summary_row


def prepare_workload_template_sheet(sheet) -> None:
    data_start, _, _ = find_workload_block(sheet)
    sample_row = data_start - 1
    if sample_row <= 1:
        return

    sample_label = normalize_text(sheet.cell(row=sample_row, column=1).value)
    sample_detail = normalize_text(sheet.cell(row=sample_row, column=2).value)
    sample_numbers = [parse_number(sheet.cell(row=sample_row, column=column).value) for column in range(3, 8)]
    if sample_label and not sample_detail and not any(number is not None for number in sample_numbers):
        sheet.delete_rows(sample_row)


def copy_row_style(sheet, source_row: int, target_row: int) -> None:
    for column in range(1, sheet.max_column + 1):
        source = sheet.cell(row=source_row, column=column)
        target = sheet.cell(row=target_row, column=column)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.font:
            target.font = copy.copy(source.font)
        if source.fill:
            target.fill = copy.copy(source.fill)
        if source.border:
            target.border = copy.copy(source.border)
    if source_row in sheet.row_dimensions:
        sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height


SOLID_SIDE = Side(style="thin", color="000000")
WORKLOAD_CONTEXT_FILL = PatternFill(
    fill_type="solid",
    fgColor=Color(theme=7, tint=0.7999816888943144),
)
WORKLOAD_A_WIDTH = 26.125
WORKLOAD_B_WIDTH = 6.63
WORKLOAD_NOTE_HEIGHT = 17.1
AMOUNT_COLUMN_WIDTHS = {
    "A": 13.0,
    "B": 10.125,
    "C": 7.0,
    "D": 5.625,
    "E": 9.125,
    "F": 15.0,
    "G": 9.125,
    "H": 13.875,
    "I": 8.0,
    "J": 13.0,
    "K": 8.0,
    "L": 13.0,
    "M": 7.125,
    "N": 13.0,
    "O": 23.25,
    "P": 11.375,
    "Q": 12.75,
    "R": 13.0,
}
AMOUNT_MIN_DETAIL_START_ROW = 29
AMOUNT_LABOR_ROWS = [
    ("기  술  사", "H", "$P$5"),
    ("특급기술자", "I", "$P$6"),
    ("고급기술자", "J", "$P$7"),
    ("중급기술자", "K", "$P$8"),
    ("초급기술자", "L", "$P$9"),
]
AMOUNT_RATE_LABELS = [
    "기술사",
    "특급기술자",
    "고급기술자",
    "중급기술자",
    "초급기술자",
    "고급숙련기술자",
    "중급숙련기술자",
    "초급숙련기술자",
]
AMOUNT_TOP_MERGES = [
    "A2:A3",
    "B2:B3",
    "C2:C3",
    "D2:D3",
    "E2:F2",
    "G2:H2",
    "I2:J2",
    "K2:L2",
    "M2:M3",
    "A4:B4",
    "A5:B5",
    "A6:B6",
    "A7:B7",
    "A10:B10",
    "A11:B11",
    "A12:B12",
    "A13:B13",
]


def set_border_side(cell, left=None, right=None, top=None, bottom=None) -> None:
    border = cell.border
    cell.border = Border(
        left=left if left is not None else border.left,
        right=right if right is not None else border.right,
        top=top if top is not None else border.top,
        bottom=bottom if bottom is not None else border.bottom,
        diagonal=border.diagonal,
        diagonal_direction=border.diagonal_direction,
        diagonalUp=border.diagonalUp,
        diagonalDown=border.diagonalDown,
        outline=border.outline,
        vertical=border.vertical,
        horizontal=border.horizontal,
    )


def set_font_style(cell, *, bold: bool | None = None, color: str | None = None, size: float | None = None) -> None:
    font = copy.copy(cell.font)
    if bold is not None:
        font.bold = bold
    if color is not None:
        font.color = color
    if size is not None:
        font.sz = size
    cell.font = font


def apply_solid_bottom(sheet, row: int, start_column: int = 1, end_column: int = 15) -> None:
    for column in range(start_column, min(end_column, sheet.max_column) + 1):
        set_border_side(sheet.cell(row=row, column=column), bottom=SOLID_SIDE)


def apply_solid_top(sheet, row: int, start_column: int = 1, end_column: int = 15) -> None:
    for column in range(start_column, min(end_column, sheet.max_column) + 1):
        set_border_side(sheet.cell(row=row, column=column), top=SOLID_SIDE)


def apply_solid_vertical(sheet, start_row: int, end_row: int, left_column: int, right_column: int) -> None:
    for row in range(start_row, end_row + 1):
        set_border_side(sheet.cell(row=row, column=left_column), right=SOLID_SIDE)
        set_border_side(sheet.cell(row=row, column=right_column), left=SOLID_SIDE)


CELL_REF_PATTERN = re.compile(r"(?<!!)(\$?)([A-Z]{1,3})(\$?)(\d+)")


def shift_formula_refs_left(formula: str) -> str:
    def replace_ref(match: re.Match[str]) -> str:
        column_index = column_index_from_string(match.group(2))
        if column_index <= 1:
            return match.group(0)
        column = get_column_letter(column_index - 1)
        return f"{match.group(1)}{column}{match.group(3)}{match.group(4)}"

    return CELL_REF_PATTERN.sub(replace_ref, formula)


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy.copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy.copy(source.alignment)
    if source.font:
        target.font = copy.copy(source.font)
    if source.fill:
        target.fill = copy.copy(source.fill)
    if source.border:
        target.border = copy.copy(source.border)


def snapshot_row_style(sheet, row: int, max_column: int = 18) -> dict[str, object]:
    cells = []
    for column in range(1, max_column + 1):
        cell = sheet.cell(row=row, column=column)
        cells.append(
            {
                "style": copy.copy(cell._style) if cell.has_style else None,
                "number_format": cell.number_format,
                "alignment": copy.copy(cell.alignment),
                "font": copy.copy(cell.font),
                "fill": copy.copy(cell.fill),
                "border": copy.copy(cell.border),
            }
        )
    return {"height": sheet.row_dimensions[row].height, "cells": cells}


def apply_row_style_snapshot(sheet, target_row: int, snapshot: dict[str, object]) -> None:
    height = snapshot.get("height")
    if height is not None:
        sheet.row_dimensions[target_row].height = height
    cells = snapshot.get("cells", [])
    if not isinstance(cells, list):
        return
    for column, style in enumerate(cells, start=1):
        if not isinstance(style, dict):
            continue
        cell = sheet.cell(row=target_row, column=column)
        if style.get("style") is not None:
            cell._style = copy.copy(style["style"])
        cell.number_format = style.get("number_format", cell.number_format)
        cell.alignment = copy.copy(style.get("alignment", cell.alignment))
        cell.font = copy.copy(style.get("font", cell.font))
        cell.fill = copy.copy(style.get("fill", cell.fill))
        cell.border = copy.copy(style.get("border", cell.border))


def find_workload_header_row(sheet) -> int:
    for row in range(1, sheet.max_row + 1):
        if normalize_text(sheet.cell(row=row, column=1).value) == "구분":
            return row
        if normalize_text(sheet.cell(row=row, column=2).value) == "구분":
            return row
    return 4


def is_major_work_item(label: object) -> bool:
    text = normalize_text(label)
    return bool(re.match(r"^\s*\d+\)", text)) or bool(re.match(r"^\s*\d+\.(?!\d)", text))


def style_workload_sheet(sheet, data_start: int, summary_row: int, header_row_count: int = 0) -> None:
    header_row = find_workload_header_row(sheet)
    label_column = 1 if normalize_text(sheet.cell(row=header_row, column=1).value) == "구분" else 2
    table_end_column = 14 if label_column == 1 else 15
    factor_column = table_end_column
    pre_boundary = (7, 8) if label_column == 1 else (8, 9)
    remark_boundary = (13, 14) if label_column == 1 else (14, 15)
    sheet.row_dimensions[summary_row].height = 24
    apply_solid_bottom(sheet, header_row, 1, table_end_column)
    apply_solid_vertical(sheet, header_row, summary_row, pre_boundary[0], pre_boundary[1])
    apply_solid_vertical(sheet, header_row, summary_row, remark_boundary[0], remark_boundary[1])

    if label_column == 1:
        sheet.column_dimensions["A"].width = WORKLOAD_A_WIDTH
        sheet.column_dimensions["B"].width = WORKLOAD_B_WIDTH
        set_border_side(sheet.cell(row=header_row, column=1), right=SOLID_SIDE, bottom=SOLID_SIDE)
        set_border_side(sheet.cell(row=header_row + 1, column=1), right=SOLID_SIDE, bottom=SOLID_SIDE)
        set_border_side(sheet.cell(row=header_row, column=2), right=SOLID_SIDE)
        set_border_side(sheet.cell(row=header_row, column=8), left=SOLID_SIDE, right=SOLID_SIDE)
        set_border_side(sheet.cell(row=header_row, column=14), left=SOLID_SIDE)
    else:
        set_border_side(sheet.cell(row=header_row, column=3), right=SOLID_SIDE)
        set_border_side(sheet.cell(row=header_row, column=9), left=SOLID_SIDE, right=SOLID_SIDE)
        set_border_side(sheet.cell(row=header_row, column=15), left=SOLID_SIDE)

    empty_fill = PatternFill(fill_type=None)
    header_fill_start = max(1, data_start - header_row_count)
    for row in range(header_fill_start, summary_row):
        for column in range(1, table_end_column + 1):
            sheet.cell(row=row, column=column).fill = copy.copy(empty_fill)

    for column in range(1, table_end_column + 1):
        sheet.cell(row=summary_row, column=column).fill = copy.copy(WORKLOAD_CONTEXT_FILL)

    for row in range(data_start, summary_row):
        label_cell = sheet.cell(row=row, column=label_column)
        label = label_cell.value
        if is_major_work_item(label):
            apply_solid_top(sheet, row, 1, table_end_column)
            set_font_style(label_cell, bold=True)

        factor_cell = sheet.cell(row=row, column=factor_column)
        if factor_cell.value not in (None, ""):
            set_font_style(factor_cell, bold=True, color="FF0000")

    for row in range(summary_row + 1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = WORKLOAD_NOTE_HEIGHT
        for column in range(1, table_end_column + 1):
            cell = sheet.cell(row=row, column=column)
            cell.border = Border()
            set_font_style(cell, bold=False)

    sheet.print_area = f"A1:{get_column_letter(table_end_column)}{sheet.max_row}"
    sheet.sheet_view.view = "pageBreakPreview"
    sheet.sheet_view.zoomScale = 100
    sheet.sheet_view.zoomScaleNormal = 100
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = 9


def delete_workload_first_column(sheet) -> None:
    header_row = find_workload_header_row(sheet)
    title = sheet.cell(row=1, column=1).value
    division_label = sheet.cell(row=header_row, column=1).value or "구분"

    for merged_range in list(sheet.merged_cells.ranges):
        sheet.unmerge_cells(str(merged_range))

    for cell_row in (header_row, header_row + 1):
        copy_cell_style(sheet.cell(row=cell_row, column=1), sheet.cell(row=cell_row, column=2))
    sheet.cell(row=header_row, column=2).value = division_label
    sheet.cell(row=header_row + 1, column=2).value = None

    sheet.delete_cols(1)

    if title:
        sheet.cell(row=1, column=1).value = title
    sheet.cell(row=header_row, column=1).value = division_label

    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = shift_formula_refs_left(cell.value)

    if sheet.max_column >= 14:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    sheet.merge_cells(start_row=header_row, start_column=1, end_row=header_row + 1, end_column=1)
    sheet.merge_cells(start_row=header_row, start_column=2, end_row=header_row, end_column=7)
    sheet.merge_cells(start_row=header_row, start_column=8, end_row=header_row, end_column=13)


def factor_formula(symbol: str) -> str | int | float:
    compact = normalize_text(symbol)
    if "①" in compact or compact == "1":
        return "=$G$3"
    if "②" in compact or compact == "2":
        return 1
    number = parse_number(compact)
    return 1 if number is None else number


def excel_number(value: float) -> str:
    return format(float(value), ".12g")


def excel_rounddown(value: float, digits: int = 0) -> float:
    factor = 10 ** int(digits)
    if digits >= 0:
        return math.floor(float(value) * factor) / factor
    factor = 10 ** abs(int(digits))
    return math.floor(float(value) / factor) * factor


def evaluate_factor_formula(formula: str, area: float) -> float:
    text = normalize_text(formula)
    if text.startswith("="):
        text = text[1:].strip()
    if not text:
        raise ValueError("직접 보정계수 산식이 비어 있습니다.")
    number = parse_number(text)
    if number is not None and re.fullmatch(r"-?\d+(?:\.\d+)?", text.replace(",", "")):
        return number

    unit_area = 0.1 if area < 1000 else round(area / 10000, 2)
    expression = text.replace("^", "**")
    names = {
        "B2": float(area),
        "D2": float(unit_area),
        "H3": 1.0,
        "ROUND": round,
        "round": round,
        "INT": lambda value: math.floor(float(value)),
        "int": lambda value: math.floor(float(value)),
        "IF": lambda condition, true_value, false_value: true_value if condition else false_value,
        "if": lambda condition, true_value, false_value: true_value if condition else false_value,
        "ROUNDDOWN": excel_rounddown,
        "rounddown": excel_rounddown,
        "ABS": abs,
        "abs": abs,
    }
    try:
        return float(eval(expression, {"__builtins__": {}}, names))
    except Exception as exc:
        raise ValueError(f"직접 보정계수 산식을 계산하지 못했습니다: {formula}") from exc


def factor_formula_for_excel(formula: str | None, numeric_value: float | None) -> str | None:
    if formula:
        expression = formula.strip()
        if expression.startswith("="):
            expression = expression[1:].strip()
        if expression:
            return f"=ROUND(({expression})*H3,3)"
    if numeric_value is not None:
        return f"=ROUND({excel_number(numeric_value)}*H3,3)"
    return None


def multiline_link_formula(refs: list[str]) -> str | None:
    if not refs:
        return None
    return "=" + "&CHAR(10)&".join(refs)


def apply_pdf_context_links(sheet, result: ExtractResult, data_start: int, summary_row: int) -> None:
    note_refs = result.raw_sheet_refs.get("notes", [])

    if note_refs:
        note_start = summary_row + 1
        for offset, ref in enumerate(note_refs):
            row = note_start + offset
            if row > sheet.max_row:
                copy_row_style(sheet, summary_row, row)
            note_cell = sheet.cell(row=row, column=2)
            note_cell.value = f"={ref}"
            note_cell.alignment = Alignment(wrap_text=False, vertical="center")
            for column in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=column)
                cell.border = Border()
                set_font_style(cell, bold=False)
            for column in list(range(3, 8)) + list(range(9, 14)) + [15]:
                sheet.cell(row=row, column=column).value = None


def write_header_rows(sheet, result: ExtractResult, data_start: int) -> int:
    header_refs = result.raw_sheet_refs.get("headers", [])
    if not header_refs:
        return 0

    sheet.insert_rows(data_start, amount=len(header_refs))
    for offset, ref in enumerate(header_refs):
        row = data_start + offset
        copy_row_style(sheet, data_start + len(header_refs), row)
        for column in list(range(2, 8)) + list(range(9, 14)) + [15]:
            sheet.cell(row=row, column=column).value = None
        header_cell = sheet.cell(row=row, column=2)
        header_cell.value = f"={ref}"
        header_cell.alignment = copy.copy(header_cell.alignment)
        header_cell.alignment = Alignment(wrap_text=False, vertical="center")
    return len(header_refs)


def fill_workload_sheet(
    sheet,
    result: ExtractResult,
    area: float | None,
    factor_override: float | None = None,
    factor_formula_override: str | None = None,
    ratio: float = 1.0,
) -> int:
    prepare_workload_template_sheet(sheet)
    data_start, data_end, summary_row = find_workload_block(sheet)
    inserted_header_rows = write_header_rows(sheet, result, data_start)
    data_start += inserted_header_rows
    data_end += inserted_header_rows
    summary_row += inserted_header_rows
    original_capacity = data_end - data_start + 1
    display_row_count = len(result.items) + sum(len(item.sub_items) for item in result.items)

    if display_row_count > original_capacity:
        extra = display_row_count - original_capacity
        sheet.insert_rows(summary_row, amount=extra)
        for row in range(summary_row, summary_row + extra):
            copy_row_style(sheet, data_end, row)
        summary_row += extra
        data_end += extra

    if area is not None:
        sheet["B2"] = area
    sheet["H3"] = ratio
    factor_cell_formula = factor_formula_for_excel(factor_formula_override, factor_override)
    if factor_cell_formula is not None:
        sheet["G3"] = factor_cell_formula
    sheet["A1"] = f"<{result.pdf_path.stem} 직접인력 소요작업량 산정>"
    sheet["G2"] = None
    sheet["Q2"] = None

    adjusted_columns = ["I", "J", "K", "L", "M"]
    raw_columns = ["C", "D", "E", "F", "G"]

    result.amount_entries = []
    row = data_start
    for item in result.items:
        item_entry: dict[str, object] = {
            "label": item.label,
            "row": row,
            "has_values": bool(not item.sub_items and not item.display_only),
            "children": [],
        }
        result.amount_entries.append(item_entry)
        label_cell = sheet.cell(row=row, column=2)
        label_cell.value = item.label
        label_cell.alignment = copy.copy(label_cell.alignment)
        label_cell.alignment = Alignment(wrap_text=True, vertical="center")
        if item.sub_items or item.display_only:
            for column in list(range(3, 8)) + list(range(9, 14)) + [15]:
                sheet.cell(row=row, column=column).value = None
        else:
            for column_offset, value in enumerate(item.values, start=3):
                sheet.cell(row=row, column=column_offset).value = value
            for target_column, source_column in zip(adjusted_columns, raw_columns, strict=True):
                sheet[f"{target_column}{row}"] = f"=ROUND({source_column}{row}*O{row},2)"
            sheet[f"O{row}"] = factor_formula(item.factor_symbol)
        row += 1

        for child in item.sub_items:
            child_entry: dict[str, object] = {
                "label": child.label,
                "row": row,
                "has_values": bool(not child.display_only),
                "children": [],
            }
            children = item_entry["children"]
            if isinstance(children, list):
                children.append(child_entry)
            detail_cell = sheet.cell(row=row, column=2)
            detail_cell.value = child.label
            detail_cell.alignment = copy.copy(detail_cell.alignment)
            detail_cell.alignment = Alignment(wrap_text=True, vertical="center")
            if child.display_only:
                for column in list(range(3, 8)) + list(range(9, 14)) + [15]:
                    sheet.cell(row=row, column=column).value = None
            else:
                for column_offset, value in enumerate(child.values, start=3):
                    sheet.cell(row=row, column=column_offset).value = value
                for target_column, source_column in zip(adjusted_columns, raw_columns, strict=True):
                    sheet[f"{target_column}{row}"] = f"=ROUND({source_column}{row}*O{row},2)"
                sheet[f"O{row}"] = factor_formula(child.factor_symbol)
            row += 1

    for row in range(data_start + display_row_count, data_end + 1):
        for column in list(range(2, 8)) + list(range(9, 14)) + [15]:
            sheet.cell(row=row, column=column).value = None

    last_data_row = data_start + display_row_count - 1
    for column in ["C", "D", "E", "F", "G", "I", "J", "K", "L", "M"]:
        sheet[f"{column}{summary_row}"] = f"=SUM({column}{data_start}:{column}{last_data_row})"

    apply_pdf_context_links(sheet, result, data_start, summary_row)
    delete_workload_first_column(sheet)
    style_workload_sheet(sheet, data_start, summary_row, inserted_header_rows)
    return summary_row


def amount_entry_has_values(entry: dict[str, object]) -> bool:
    return bool(entry.get("has_values"))


def amount_entry_children(entry: dict[str, object]) -> list[dict[str, object]]:
    children = entry.get("children", [])
    return children if isinstance(children, list) else []


def amount_entry_row(entry: dict[str, object]) -> int:
    return int(entry.get("row", 0))


def amount_ref(workload_sheet_name: str, column: str, row: int) -> str:
    return f"={quote_sheet_ref(workload_sheet_name)}{column}{row}"


def amount_add_formula(rows: list[int]) -> str | int:
    if not rows:
        return 0
    return "=" + "+".join(f"H{row}" for row in rows)


def amount_sum_formula(rows: list[int]) -> str | int:
    if not rows:
        return 0
    if len(rows) > 1 and rows == list(range(rows[0], rows[-1] + 1)):
        return f"=SUM(H{rows[0]}:H{rows[-1]})"
    return amount_add_formula(rows)


def find_amount_blank_style_row(sheet) -> int:
    for row in range(16, sheet.max_row + 1):
        if all(sheet.cell(row=row, column=column).value is None for column in range(1, 14)):
            return row
    return 24


def clear_amount_row(sheet, row: int, max_column: int = 18) -> None:
    for column in range(1, max_column + 1):
        sheet.cell(row=row, column=column).value = None


def prepare_amount_row(sheet, row: int, style_snapshot: dict[str, object], height: float = 18.0) -> None:
    apply_row_style_snapshot(sheet, row, style_snapshot)
    clear_amount_row(sheet, row)
    sheet.row_dimensions[row].height = height


def configure_amount_static_area(
    sheet,
    title: str,
    row_styles: dict[str, dict[str, object]],
    labor_rates: list[float | None] | None = None,
) -> None:
    for column, width in AMOUNT_COLUMN_WIDTHS.items():
        sheet.column_dimensions[column].width = width

    rate_values = [sheet.cell(row=row, column=16).value for row in range(5, 13)]
    if not any(value is not None for value in rate_values):
        rate_values = [sheet.cell(row=row, column=16).value for row in range(15, 23)]
    while len(rate_values) < len(AMOUNT_RATE_LABELS):
        rate_values.append(None)
    if labor_rates:
        for index, value in enumerate(labor_rates[: len(AMOUNT_RATE_LABELS)]):
            if value is not None:
                rate_values[index] = value
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row <= 15:
            sheet.unmerge_cells(str(merged_range))

    for row_index in range(1, 16):
        for column in range(1, 19):
            sheet.cell(row=row_index, column=column).value = None

    for merged_range in AMOUNT_TOP_MERGES:
        sheet.merge_cells(merged_range)

    sheet["A1"] = title
    sheet["A5"] = f"  1. {title_without_leading_number(title)}"
    sheet["A2"] = "종   별"
    sheet["B2"] = "규 격"
    sheet["C2"] = "수  량"
    sheet["D2"] = "단위"
    sheet["E2"] = "총      계"
    sheet["G2"] = "노     무     비"
    sheet["I2"] = "재     료     비"
    sheet["K2"] = "경          비"
    sheet["M2"] = "비 고"
    for column in ["E", "G", "I", "K"]:
        sheet[f"{column}3"] = "단 가"
        sheet[f"{get_column_letter(column_index_from_string(column) + 1)}3"] = "금 액"

    sheet["A4"] = " 가. 직접인건비"
    sheet["A5"] = f"  1. {title_without_leading_number(title)}"
    sheet["C5"] = 1
    sheet["D5"] = "식"
    sheet["F5"] = "=H5"
    sheet["H5"] = "=H16"
    sheet["A6"] = "소   계"
    sheet["F6"] = "=H6"
    sheet["H6"] = "=SUM(H5:H5)"
    sheet["A8"] = " 나. 제경비"
    sheet["B8"] = "(직접인건비)x "
    sheet["D8"] = 1.1
    sheet["F8"] = "=ROUNDDOWN(F6*D8,0)"
    sheet["A9"] = " 다. 기술료"
    sheet["B9"] = "(직접인건비+제경비)x"
    sheet["D9"] = 0.2
    sheet["F9"] = "=INT((F6+F8)*D9)"
    sheet["A10"] = "소   계"
    sheet["F10"] = "=F8+F9"
    sheet["A12"] = "합   계"
    sheet["F12"] = "=ROUNDDOWN(F6+F10,0)"
    sheet["A13"] = "적   용"
    sheet["F13"] = "=ROUNDDOWN(F12,-3)"
    sheet["M13"] = "단위절사"

    for row_index in range(4, 16):
        sheet.row_dimensions[row_index].height = 24
    apply_solid_bottom(sheet, 8, 1, 13)

    sheet["O3"] = "■ 엔지니어링 노임단가(2026년_건설)"
    sheet["O4"] = "구   분"
    sheet["P4"] = "노  무  비"
    sheet["Q4"] = "비   고"
    for offset, label in enumerate(AMOUNT_RATE_LABELS):
        row = 5 + offset
        sheet.cell(row=row, column=15).value = label
        sheet.cell(row=row, column=16).value = rate_values[offset] if offset < len(rate_values) else None
        sheet.cell(row=row, column=17).value = None

    for row in range(13, 23):
        for column in range(15, 18):
            sheet.cell(row=row, column=column).value = None

    apply_row_style_snapshot(sheet, 15, row_styles["static_blank"])
    clear_amount_row(sheet, 15)


def amount_groups(entries: list[dict[str, object]]) -> list[dict[str, object]]:
    groups: list[dict[str, object]] = []
    for entry in entries:
        children = [child for child in amount_entry_children(entry) if amount_entry_has_values(child)]
        if children:
            groups.append({"parent": entry, "tasks": children})
        elif amount_entry_has_values(entry):
            groups.append({"parent": None, "tasks": [entry]})
    return groups


def write_amount_task_block(
    sheet,
    start_row: int,
    task: dict[str, object],
    workload_sheet_name: str,
    row_styles: dict[str, dict[str, object]],
) -> int:
    prepare_amount_row(sheet, start_row, row_styles["detail_task"], 18.0)
    workload_row = amount_entry_row(task)
    sheet.cell(row=start_row, column=1).value = amount_ref(workload_sheet_name, "A", workload_row)
    labor_start = start_row + 1
    labor_end = start_row + len(AMOUNT_LABOR_ROWS)
    sheet.cell(row=start_row, column=8).value = f"=SUM(H{labor_start}:H{labor_end})"

    for offset, (label, workload_column, rate_ref) in enumerate(AMOUNT_LABOR_ROWS):
        row = labor_start + offset
        style_key = f"labor{offset}"
        prepare_amount_row(sheet, row, row_styles[style_key], 18.0)
        sheet.cell(row=row, column=1).value = label
        sheet.cell(row=row, column=3).value = amount_ref(workload_sheet_name, workload_column, workload_row)
        sheet.cell(row=row, column=4).value = "인"
        sheet.cell(row=row, column=7).value = f"={rate_ref}"
        sheet.cell(row=row, column=8).value = f"=ROUNDDOWN(G{row}*C{row},0)"
    return start_row


def amount_summary_row_count(groups: list[dict[str, object]]) -> int:
    count = 1
    for group in groups:
        parent = group.get("parent")
        tasks = group.get("tasks", [])
        if not isinstance(tasks, list):
            tasks = []
        count += len(tasks)
        if isinstance(parent, dict):
            count += 1
    return count


def rebuild_amount_dynamic_area(sheet, workload_sheet_name: str, entries: list[dict[str, object]], row_styles: dict[str, dict[str, object]]) -> int:
    groups = amount_groups(entries)
    summary_start = 14
    summary_count = amount_summary_row_count(groups)
    detail_start = max(AMOUNT_MIN_DETAIL_START_ROW, summary_start + summary_count)

    prepare_amount_row(sheet, detail_start, row_styles["detail_overall"], 24.0)
    sheet.cell(row=detail_start, column=1).value = "=A5"
    current_row = detail_start + 1
    detail_top_rows: list[int] = []
    detail_rows: dict[int, int] = {}

    for group_index, group in enumerate(groups):
        if group_index > 0:
            prepare_amount_row(sheet, current_row, row_styles["blank"], 18.0)
            current_row += 1

        parent = group.get("parent")
        tasks = group.get("tasks", [])
        if not isinstance(tasks, list):
            tasks = []

        if isinstance(parent, dict):
            parent_row = current_row
            prepare_amount_row(sheet, parent_row, row_styles["detail_parent"], 18.0)
            detail_rows[amount_entry_row(parent)] = parent_row
            sheet.cell(row=parent_row, column=1).value = amount_ref(workload_sheet_name, "A", amount_entry_row(parent))
            current_row += 1
            child_total_rows: list[int] = []
            for task in tasks:
                if isinstance(task, dict):
                    task_row = write_amount_task_block(sheet, current_row, task, workload_sheet_name, row_styles)
                    detail_rows[amount_entry_row(task)] = task_row
                    child_total_rows.append(task_row)
                    current_row += 1 + len(AMOUNT_LABOR_ROWS)
            sheet.cell(row=parent_row, column=8).value = amount_add_formula(child_total_rows)
            detail_top_rows.append(parent_row)
        else:
            for task in tasks:
                if isinstance(task, dict):
                    task_row = write_amount_task_block(sheet, current_row, task, workload_sheet_name, row_styles)
                    detail_rows[amount_entry_row(task)] = task_row
                    detail_top_rows.append(task_row)
                    current_row += 1 + len(AMOUNT_LABOR_ROWS)

    sheet.cell(row=detail_start, column=8).value = amount_add_formula(detail_top_rows)

    prepare_amount_row(sheet, summary_start, row_styles["summary_overall"], 24.0)
    sheet.cell(row=summary_start, column=1).value = "=A5"
    summary_row = summary_start + 1
    summary_top_rows: list[int] = []

    for group in groups:
        parent = group.get("parent")
        tasks = group.get("tasks", [])
        if not isinstance(tasks, list):
            tasks = []

        if isinstance(parent, dict):
            parent_summary_row = summary_row
            prepare_amount_row(sheet, parent_summary_row, row_styles["summary_parent"], 18.0)
            detail_parent_row = detail_rows.get(amount_entry_row(parent))
            sheet.cell(row=parent_summary_row, column=1).value = f"=A{detail_parent_row}" if detail_parent_row else amount_ref(workload_sheet_name, "A", amount_entry_row(parent))
            summary_row += 1
            child_summary_rows: list[int] = []
            for task in tasks:
                if not isinstance(task, dict):
                    continue
                child_summary_row = summary_row
                prepare_amount_row(sheet, child_summary_row, row_styles["summary_child"], 18.0)
                detail_task_row = detail_rows.get(amount_entry_row(task))
                sheet.cell(row=child_summary_row, column=1).value = f"=A{detail_task_row}" if detail_task_row else amount_ref(workload_sheet_name, "A", amount_entry_row(task))
                sheet.cell(row=child_summary_row, column=6).value = f"=H{child_summary_row}"
                sheet.cell(row=child_summary_row, column=8).value = f"=H{detail_task_row}" if detail_task_row else 0
                child_summary_rows.append(child_summary_row)
                summary_row += 1
            sheet.cell(row=parent_summary_row, column=8).value = amount_sum_formula(child_summary_rows)
            summary_top_rows.append(parent_summary_row)
        else:
            for task in tasks:
                if not isinstance(task, dict):
                    continue
                task_summary_row = summary_row
                prepare_amount_row(sheet, task_summary_row, row_styles["summary_child"], 18.0)
                detail_task_row = detail_rows.get(amount_entry_row(task))
                sheet.cell(row=task_summary_row, column=1).value = f"=A{detail_task_row}" if detail_task_row else amount_ref(workload_sheet_name, "A", amount_entry_row(task))
                sheet.cell(row=task_summary_row, column=6).value = f"=H{task_summary_row}"
                sheet.cell(row=task_summary_row, column=8).value = f"=H{detail_task_row}" if detail_task_row else 0
                summary_top_rows.append(task_summary_row)
                summary_row += 1

    sheet.cell(row=summary_start, column=8).value = amount_add_formula(summary_top_rows)

    for row in range(summary_row, detail_start):
        prepare_amount_row(sheet, row, row_styles["blank"], 18.0)

    for _ in range(1):
        prepare_amount_row(sheet, current_row, row_styles["blank"], 18.0)
        current_row += 1

    last_row = current_row - 1
    if sheet.max_row > last_row:
        sheet.delete_rows(last_row + 1, sheet.max_row - last_row)
    return last_row


def update_amount_sheet(
    sheet,
    old_workload_name: str,
    new_workload_name: str,
    summary_row: int,
    template_summary_row: int,
    title: str,
    result: ExtractResult,
    labor_rates: list[float | None] | None = None,
) -> None:
    _ = (old_workload_name, summary_row, template_summary_row)
    row_styles = {
        "static_blank": snapshot_row_style(sheet, 15 if sheet.max_row >= 15 else find_amount_blank_style_row(sheet)),
        "summary_overall": snapshot_row_style(sheet, 16 if sheet.max_row >= 16 else 4),
        "summary_parent": snapshot_row_style(sheet, 17 if sheet.max_row >= 17 else 4),
        "summary_child": snapshot_row_style(sheet, 18 if sheet.max_row >= 18 else 4),
        "detail_overall": snapshot_row_style(sheet, 29 if sheet.max_row >= 29 else 15),
        "detail_parent": snapshot_row_style(sheet, 30 if sheet.max_row >= 30 else 16),
        "detail_task": snapshot_row_style(sheet, 31 if sheet.max_row >= 31 else 17),
        "labor0": snapshot_row_style(sheet, 32 if sheet.max_row >= 32 else 18),
        "labor1": snapshot_row_style(sheet, 33 if sheet.max_row >= 33 else 19),
        "labor2": snapshot_row_style(sheet, 34 if sheet.max_row >= 34 else 20),
        "labor3": snapshot_row_style(sheet, 35 if sheet.max_row >= 35 else 21),
        "labor4": snapshot_row_style(sheet, 36 if sheet.max_row >= 36 else 22),
        "blank": snapshot_row_style(sheet, 28 if sheet.max_row >= 28 else find_amount_blank_style_row(sheet)),
    }
    configure_amount_static_area(sheet, title, row_styles, labor_rates=labor_rates)
    last_row = rebuild_amount_dynamic_area(sheet, new_workload_name, result.amount_entries, row_styles)

    sheet.print_area = f"A1:M{last_row}"
    sheet.sheet_view.view = "pageBreakPreview"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = 9
    sheet.page_setup.scale = 90
    sheet.page_setup.fitToWidth = None
    sheet.page_setup.fitToHeight = None


def write_raw_sheet(workbook: openpyxl.Workbook, result: ExtractResult, sheet_name: str) -> None:
    sheet = workbook.create_sheet(sheet_name)
    refs: dict[str, list[str]] = {"headers": [], "notes": []}
    sheet["A1"] = "출처 PDF"
    sheet["B1"] = str(result.pdf_path)
    sheet["A2"] = "쪽수"
    sheet["B2"] = result.page_count
    sheet["A3"] = "추출일시"
    sheet["B3"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(bold=True)
    for cell in sheet[1] + sheet[2] + sheet[3]:
        cell.alignment = Alignment(vertical="center")
    for cell in ("A1", "A2", "A3"):
        sheet[cell].font = header_font
        sheet[cell].fill = header_fill

    row_cursor = 5
    if result.header_lines:
        sheet.cell(row=row_cursor, column=1).value = "상단 머릿글"
        sheet.cell(row=row_cursor, column=1).font = header_font
        row_cursor += 1
        for line in result.header_lines:
            cell = sheet.cell(row=row_cursor, column=1)
            cell.value = line
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            refs["headers"].append(f"{quote_sheet_ref(sheet.title)}{cell.coordinate}")
            row_cursor += 1
        row_cursor += 1

    for table_index, table in enumerate(result.raw_tables, start=1):
        sheet.cell(row=row_cursor, column=1).value = f"표 {table_index}"
        sheet.cell(row=row_cursor, column=1).font = header_font
        row_cursor += 1
        for row in table:
            for column_index, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_cursor, column=column_index)
                cell.value = value
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_cursor += 1
        row_cursor += 1

    if result.note_lines:
        sheet.cell(row=row_cursor, column=1).value = "주기"
        set_font_style(sheet.cell(row=row_cursor, column=1), bold=False)
        row_cursor += 1
        for line in result.note_lines:
            cell = sheet.cell(row=row_cursor, column=1)
            cell.value = line
            cell.alignment = Alignment(wrap_text=False, vertical="top")
            set_font_style(cell, bold=False)
            sheet.row_dimensions[row_cursor].height = WORKLOAD_NOTE_HEIGHT
            for column in range(1, sheet.max_column + 1):
                sheet.cell(row=row_cursor, column=column).border = Border()
            refs["notes"].append(f"{quote_sheet_ref(sheet.title)}{cell.coordinate}")
            row_cursor += 1

    result.raw_sheet_refs = refs
    autosize_columns(sheet)


def autosize_columns(sheet, max_width: int = 45) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            value = normalize_text(cell.value)
            if value:
                width = max(width, min(max_width, len(value) + 2))
        sheet.column_dimensions[column_letter].width = width


def write_log_sheet(workbook: openpyxl.Workbook, rows: list[list[object]]) -> None:
    name = unique_sheet_name(workbook, "_변환기록")
    sheet = workbook.create_sheet(name, 0)
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
    autosize_columns(sheet, max_width=80)


def configure_recalculation(workbook: openpyxl.Workbook) -> None:
    calculation = workbook.calculation
    calculation.calcMode = "auto"
    calculation.fullCalcOnLoad = True
    calculation.forceFullCalc = True


def build_workbook(
    template_path: Path,
    pdf_paths: Iterable[Path],
    output_path: Path,
    area: float | None = None,
    factor_override: float | None = None,
    factor_formula_override: str | None = None,
    ratio: float = 1.0,
    labor_rates: list[float | None] | None = None,
    keep_templates: bool = False,
) -> list[ExtractResult]:
    pdf_paths = [Path(path) for path in pdf_paths]
    if not pdf_paths:
        raise ValueError("변환할 PDF 파일을 선택해야 합니다.")
    if not template_path.exists():
        raise FileNotFoundError(f"템플릿 엑셀 파일이 없습니다: {template_path}")
    for pdf_path in pdf_paths:
        if not pdf_path.exists():
            raise FileNotFoundError(f"PDF 파일이 없습니다: {pdf_path}")

    workbook = openpyxl.load_workbook(template_path)
    if AMOUNT_TEMPLATE_SHEET not in workbook.sheetnames or WORKLOAD_TEMPLATE_SHEET not in workbook.sheetnames:
        raise ValueError(
            f"템플릿에는 '{AMOUNT_TEMPLATE_SHEET}', '{WORKLOAD_TEMPLATE_SHEET}' 시트가 필요합니다."
        )

    amount_template = workbook[AMOUNT_TEMPLATE_SHEET]
    workload_template = workbook[WORKLOAD_TEMPLATE_SHEET]
    _, _, template_summary_row = find_workload_block(workload_template)

    log_rows: list[list[object]] = [
        ["항목", "내용"],
        ["템플릿", str(template_path)],
        ["대상면적(㎡)", "" if area is None else area],
        ["보정계수 직접입력", "" if factor_override is None else factor_override],
        ["보정계수 산식", "" if factor_formula_override is None else factor_formula_override],
        ["비율", ratio],
        ["변환일시", dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]

    results: list[ExtractResult] = []
    for pdf_path in pdf_paths:
        result = extract_pdf(pdf_path)
        results.append(result)
        base = sanitize_sheet_base(pdf_path.stem)

        raw_name = make_sheet_name(workbook, base, "_PDF")
        workload_name = make_sheet_name(workbook, base, "_작업량")
        amount_name = make_sheet_name(workbook, base, "_금액")

        amount_sheet = workbook.copy_worksheet(amount_template)
        amount_sheet.title = amount_name
        workload_sheet = workbook.copy_worksheet(workload_template)
        workload_sheet.title = workload_name

        write_raw_sheet(workbook, result, raw_name)
        summary_row = fill_workload_sheet(
            workload_sheet,
            result,
            area,
            factor_override=factor_override,
            factor_formula_override=factor_formula_override,
            ratio=ratio,
        )
        update_amount_sheet(
            amount_sheet,
            old_workload_name=WORKLOAD_TEMPLATE_SHEET,
            new_workload_name=workload_name,
            summary_row=summary_row,
            template_summary_row=template_summary_row,
            title=pdf_path.stem,
            result=result,
            labor_rates=labor_rates,
        )

        log_rows.append([])
        log_rows.append(["PDF", str(pdf_path)])
        log_rows.append(["추출 표 수", len(result.raw_tables)])
        log_rows.append(["반영 업무 행 수", len(result.items)])
        if result.warnings:
            log_rows.append(["주의", " / ".join(result.warnings)])

    if not keep_templates:
        workbook.remove(amount_template)
        workbook.remove(workload_template)

    write_log_sheet(workbook, log_rows)
    configure_recalculation(workbook)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return results


def legacy_launch_gui() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    root = tk.Tk()
    root.title("PDF 표 엑셀 산식 변환기")
    root.geometry("760x440")
    root.minsize(720, 420)

    template_var = tk.StringVar(value=str(DEFAULT_TEMPLATE if DEFAULT_TEMPLATE.exists() else ""))
    output_var = tk.StringVar(value=str(DEFAULT_OUTPUT))
    area_var = tk.StringVar(value="10000")
    status_var = tk.StringVar(value="PDF와 템플릿 엑셀을 선택한 뒤 변환을 실행하세요.")
    pdf_paths: list[str] = [str(DEFAULT_PDF)] if DEFAULT_PDF.exists() else []

    def refresh_pdf_list() -> None:
        pdf_list.delete(0, tk.END)
        for path in pdf_paths:
            pdf_list.insert(tk.END, path)

    def choose_template() -> None:
        path = filedialog.askopenfilename(
            title="템플릿 엑셀 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            template_var.set(path)

    def choose_pdfs() -> None:
        paths = filedialog.askopenfilenames(
            title="변환할 PDF 선택",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if paths:
            pdf_paths.clear()
            pdf_paths.extend(paths)
            refresh_pdf_list()

    def choose_output() -> None:
        path = filedialog.asksaveasfilename(
            title="저장할 엑셀 파일",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=Path(output_var.get()).name,
        )
        if path:
            output_var.set(path)

    def run_conversion() -> None:
        try:
            area_text = area_var.get().strip().replace(",", "")
            area = float(area_text) if area_text else None
            status_var.set("변환 중입니다. PDF 표를 읽고 엑셀 수식을 연결하고 있습니다...")
            root.update_idletasks()
            results = build_workbook(
                template_path=Path(template_var.get()),
                pdf_paths=[Path(path) for path in pdf_paths],
                output_path=Path(output_var.get()),
                area=area,
            )
        except Exception as exc:
            status_var.set("변환 실패")
            messagebox.showerror("변환 실패", str(exc))
            return

        count = sum(len(result.items) for result in results)
        status_var.set(f"완료: 업무 행 {count}개를 수식 엑셀로 반영했습니다.")
        messagebox.showinfo("변환 완료", f"저장했습니다:\n{output_var.get()}")

    outer = ttk.Frame(root, padding=16)
    outer.pack(fill=tk.BOTH, expand=True)
    outer.columnconfigure(1, weight=1)
    outer.rowconfigure(3, weight=1)

    ttk.Label(outer, text="템플릿 엑셀").grid(row=0, column=0, sticky="w", pady=4)
    ttk.Entry(outer, textvariable=template_var).grid(row=0, column=1, sticky="ew", padx=8, pady=4)
    ttk.Button(outer, text="찾기", command=choose_template).grid(row=0, column=2, sticky="ew", pady=4)

    ttk.Label(outer, text="저장 파일").grid(row=1, column=0, sticky="w", pady=4)
    ttk.Entry(outer, textvariable=output_var).grid(row=1, column=1, sticky="ew", padx=8, pady=4)
    ttk.Button(outer, text="저장 위치", command=choose_output).grid(row=1, column=2, sticky="ew", pady=4)

    ttk.Label(outer, text="대상면적(㎡)").grid(row=2, column=0, sticky="w", pady=4)
    ttk.Entry(outer, textvariable=area_var, width=16).grid(row=2, column=1, sticky="w", padx=8, pady=4)
    ttk.Button(outer, text="PDF 선택", command=choose_pdfs).grid(row=2, column=2, sticky="ew", pady=4)

    list_frame = ttk.LabelFrame(outer, text="변환할 PDF")
    list_frame.grid(row=3, column=0, columnspan=3, sticky="nsew", pady=10)
    list_frame.rowconfigure(0, weight=1)
    list_frame.columnconfigure(0, weight=1)

    pdf_list = tk.Listbox(list_frame, height=8)
    pdf_list.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)
    scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=pdf_list.yview)
    scrollbar.grid(row=0, column=1, sticky="ns", pady=8)
    pdf_list.configure(yscrollcommand=scrollbar.set)
    refresh_pdf_list()

    ttk.Label(outer, textvariable=status_var).grid(row=4, column=0, columnspan=2, sticky="w", pady=4)
    ttk.Button(outer, text="변환 실행", command=run_conversion).grid(row=4, column=2, sticky="ew", pady=4)

    root.mainloop()


def launch_gui(start_watch: bool = False) -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    def default_watch_folder() -> str:
        if DEFAULT_PDF.exists():
            return str(DEFAULT_PDF.parent)
        if DEFAULT_TEMPLATE.exists():
            return str(DEFAULT_TEMPLATE.parent)
        return str(Path.home() / "Desktop")

    def pdf_signature(paths: Iterable[str]) -> dict[str, tuple[int, int]]:
        signatures: dict[str, tuple[int, int]] = {}
        for path_text in paths:
            path = Path(path_text)
            try:
                stat = path.stat()
            except OSError:
                continue
            signatures[str(path.resolve())] = (stat.st_mtime_ns, stat.st_size)
        return signatures

    root = tk.Tk()
    root.title("PDF 표 엑셀 산식 자동 변환기")
    root.geometry("900x620")
    root.minsize(820, 560)

    template_var = tk.StringVar(value=str(DEFAULT_TEMPLATE if DEFAULT_TEMPLATE.exists() else ""))
    watch_folder_var = tk.StringVar(value=default_watch_folder())
    output_var = tk.StringVar(value=str(DEFAULT_OUTPUT))
    area_var = tk.StringVar(value="10000")
    status_var = tk.StringVar(value="대기 중: PDF 폴더를 감시하거나 수동 변환을 실행하세요.")
    watch_button_var = tk.StringVar(value="자동 감시 시작")
    pdf_paths: list[str] = [str(DEFAULT_PDF)] if DEFAULT_PDF.exists() else []
    watch_state = {
        "active": False,
        "last_signature": {},
        "tick_id": None,
        "convert_id": None,
        "busy": False,
    }

    def append_log(message: str) -> None:
        timestamp = dt.datetime.now().strftime("%H:%M:%S")
        log_box.configure(state=tk.NORMAL)
        log_box.insert(tk.END, f"[{timestamp}] {message}\n")
        log_box.see(tk.END)
        log_box.configure(state=tk.DISABLED)

    def refresh_pdf_list() -> None:
        pdf_list.delete(0, tk.END)
        for path in pdf_paths:
            pdf_list.insert(tk.END, path)

    def load_pdfs_from_folder() -> list[str]:
        folder = Path(watch_folder_var.get().strip())
        if not folder.exists():
            raise FileNotFoundError(f"감시 폴더가 없습니다: {folder}")
        if not folder.is_dir():
            raise NotADirectoryError(f"감시 폴더가 아닙니다: {folder}")
        paths = sorted(folder.glob("*.pdf"), key=lambda path: path.name.lower())
        pdf_paths.clear()
        pdf_paths.extend(str(path) for path in paths)
        refresh_pdf_list()
        return pdf_paths

    def choose_template() -> None:
        path = filedialog.askopenfilename(
            title="템플릿 엑셀 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            template_var.set(path)
            append_log(f"템플릿 선택: {path}")

    def choose_watch_folder() -> None:
        path = filedialog.askdirectory(title="PDF 감시 폴더 선택")
        if path:
            watch_folder_var.set(path)
            try:
                load_pdfs_from_folder()
            except Exception as exc:
                messagebox.showerror("폴더 확인 실패", str(exc))
                return
            append_log(f"감시 폴더 선택: {path}")

    def choose_pdfs() -> None:
        paths = filedialog.askopenfilenames(
            title="변환할 PDF 선택",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if paths:
            pdf_paths.clear()
            pdf_paths.extend(paths)
            refresh_pdf_list()
            append_log(f"수동 PDF 선택: {len(paths)}개")

    def choose_output() -> None:
        path = filedialog.asksaveasfilename(
            title="저장할 엑셀 파일",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=Path(output_var.get()).name,
        )
        if path:
            output_var.set(path)
            append_log(f"저장 파일 선택: {path}")

    def parse_area() -> float | None:
        area_text = area_var.get().strip().replace(",", "")
        return float(area_text) if area_text else None

    def run_conversion(show_popup: bool = True, reason: str = "수동 실행") -> bool:
        if watch_state["busy"]:
            append_log("이전 변환이 아직 끝나지 않아 이번 실행을 건너뜁니다.")
            return False
        try:
            watch_state["busy"] = True
            area = parse_area()
            selected_pdfs = [Path(path) for path in pdf_paths]
            status_var.set("변환 중: PDF 표를 읽고 엑셀 수식을 갱신하고 있습니다.")
            root.update_idletasks()
            results = build_workbook(
                template_path=Path(template_var.get()),
                pdf_paths=selected_pdfs,
                output_path=Path(output_var.get()),
                area=area,
            )
        except Exception as exc:
            status_var.set("변환 실패")
            append_log(f"변환 실패: {exc}")
            if show_popup:
                messagebox.showerror("변환 실패", str(exc))
            return False
        finally:
            watch_state["busy"] = False

        count = sum(len(result.items) for result in results)
        status_var.set(f"완료: 업무 행 {count}개를 반영했습니다.")
        append_log(f"{reason}: PDF {len(results)}개, 업무 행 {count}개 반영 -> {output_var.get()}")
        if show_popup:
            messagebox.showinfo("변환 완료", f"저장했습니다:\n{output_var.get()}")
        return True

    def schedule_auto_conversion(reason: str) -> None:
        convert_id = watch_state.get("convert_id")
        if convert_id is not None:
            root.after_cancel(convert_id)
        watch_state["convert_id"] = root.after(1200, lambda: auto_convert(reason))

    def auto_convert(reason: str) -> None:
        watch_state["convert_id"] = None
        if not watch_state["active"]:
            return
        current_signature = pdf_signature(pdf_paths)
        if current_signature != watch_state["last_signature"]:
            watch_state["last_signature"] = current_signature
            schedule_auto_conversion("PDF 변경 안정화 대기 후 자동 변환")
            return
        run_conversion(show_popup=False, reason=reason)

    def watch_tick() -> None:
        if not watch_state["active"]:
            return
        try:
            load_pdfs_from_folder()
            current_signature = pdf_signature(pdf_paths)
        except Exception as exc:
            status_var.set("감시 오류")
            append_log(f"감시 오류: {exc}")
            current_signature = {}

        if current_signature and current_signature != watch_state["last_signature"]:
            watch_state["last_signature"] = current_signature
            status_var.set("PDF 변경 감지: 잠시 후 엑셀을 자동 갱신합니다.")
            append_log("PDF 변경 감지")
            schedule_auto_conversion("자동 변환")
        elif not current_signature:
            status_var.set("감시 중: 폴더에 PDF가 없습니다.")

        watch_state["tick_id"] = root.after(3000, watch_tick)

    def stop_watch() -> None:
        watch_state["active"] = False
        for key in ("tick_id", "convert_id"):
            after_id = watch_state.get(key)
            if after_id is not None:
                root.after_cancel(after_id)
                watch_state[key] = None
        watch_button_var.set("자동 감시 시작")
        status_var.set("자동 감시 중지됨")
        append_log("자동 감시 중지")

    def toggle_watch() -> None:
        if watch_state["active"]:
            stop_watch()
            return
        try:
            load_pdfs_from_folder()
        except Exception as exc:
            messagebox.showerror("감시 시작 실패", str(exc))
            return
        if not pdf_paths:
            messagebox.showwarning("PDF 없음", "감시 폴더에 PDF 파일이 없습니다.")
            return
        watch_state["active"] = True
        watch_state["last_signature"] = {}
        watch_button_var.set("자동 감시 중지")
        status_var.set("자동 감시 시작: PDF가 바뀌면 결과 엑셀을 다시 저장합니다.")
        append_log("자동 감시 시작")
        watch_tick()

    def on_close() -> None:
        if watch_state["active"]:
            stop_watch()
        root.destroy()

    outer = ttk.Frame(root, padding=16)
    outer.pack(fill=tk.BOTH, expand=True)
    outer.columnconfigure(1, weight=1)
    outer.rowconfigure(5, weight=1)
    outer.rowconfigure(6, weight=1)

    ttk.Label(outer, text="템플릿 엑셀").grid(row=0, column=0, sticky="w", pady=4)
    ttk.Entry(outer, textvariable=template_var).grid(row=0, column=1, sticky="ew", padx=8, pady=4)
    ttk.Button(outer, text="찾기", command=choose_template).grid(row=0, column=2, sticky="ew", pady=4)

    ttk.Label(outer, text="PDF 감시 폴더").grid(row=1, column=0, sticky="w", pady=4)
    ttk.Entry(outer, textvariable=watch_folder_var).grid(row=1, column=1, sticky="ew", padx=8, pady=4)
    ttk.Button(outer, text="폴더 선택", command=choose_watch_folder).grid(row=1, column=2, sticky="ew", pady=4)

    ttk.Label(outer, text="저장 엑셀").grid(row=2, column=0, sticky="w", pady=4)
    ttk.Entry(outer, textvariable=output_var).grid(row=2, column=1, sticky="ew", padx=8, pady=4)
    ttk.Button(outer, text="저장 위치", command=choose_output).grid(row=2, column=2, sticky="ew", pady=4)

    ttk.Label(outer, text="대상면적(㎡)").grid(row=3, column=0, sticky="w", pady=4)
    ttk.Entry(outer, textvariable=area_var, width=16).grid(row=3, column=1, sticky="w", padx=8, pady=4)
    ttk.Button(outer, text="PDF 직접 선택", command=choose_pdfs).grid(row=3, column=2, sticky="ew", pady=4)

    button_frame = ttk.Frame(outer)
    button_frame.grid(row=4, column=0, columnspan=3, sticky="ew", pady=6)
    button_frame.columnconfigure(0, weight=1)
    button_frame.columnconfigure(1, weight=1)
    ttk.Button(button_frame, textvariable=watch_button_var, command=toggle_watch).grid(row=0, column=0, sticky="ew", padx=(0, 4))
    ttk.Button(button_frame, text="지금 한 번 변환", command=lambda: run_conversion(show_popup=True)).grid(row=0, column=1, sticky="ew", padx=(4, 0))

    list_frame = ttk.LabelFrame(outer, text="변환 대상 PDF")
    list_frame.grid(row=5, column=0, columnspan=3, sticky="nsew", pady=8)
    list_frame.rowconfigure(0, weight=1)
    list_frame.columnconfigure(0, weight=1)

    pdf_list = tk.Listbox(list_frame, height=8)
    pdf_list.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)
    scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=pdf_list.yview)
    scrollbar.grid(row=0, column=1, sticky="ns", pady=8)
    pdf_list.configure(yscrollcommand=scrollbar.set)
    refresh_pdf_list()

    log_frame = ttk.LabelFrame(outer, text="자동화 로그")
    log_frame.grid(row=6, column=0, columnspan=3, sticky="nsew", pady=8)
    log_frame.rowconfigure(0, weight=1)
    log_frame.columnconfigure(0, weight=1)
    log_box = tk.Text(log_frame, height=8, wrap="word", state=tk.DISABLED)
    log_box.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)
    log_scroll = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=log_box.yview)
    log_scroll.grid(row=0, column=1, sticky="ns", pady=8)
    log_box.configure(yscrollcommand=log_scroll.set)

    ttk.Label(outer, textvariable=status_var).grid(row=7, column=0, columnspan=3, sticky="ew", pady=4)

    append_log("프로그램 준비 완료")
    root.protocol("WM_DELETE_WINDOW", on_close)
    if start_watch:
        root.after(500, toggle_watch)
    root.mainloop()


def parse_labor_rates(value: str | None) -> list[float | None] | None:
    if value is None or not value.strip():
        return None
    rates: list[float | None] = []
    for index, raw_rate in enumerate(value.split(","), start=1):
        rate = raw_rate.strip()
        if not rate:
            rates.append(None)
            continue
        try:
            rates.append(float(rate))
        except ValueError as exc:
            raise ValueError(f"--labor-rates {index}번째 값이 숫자가 아닙니다: {raw_rate}") from exc
    return rates


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="PDF 표를 엑셀 수식 템플릿에 반영합니다.")
    parser.add_argument("pdfs", nargs="*", type=Path, help="변환할 PDF 파일")
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE, help="수식 템플릿 엑셀")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="저장할 엑셀 파일")
    parser.add_argument("--area", type=float, default=10000, help="대상면적(㎡). 빈 값은 GUI에서만 지원합니다.")
    parser.add_argument("--factor", type=float, default=None, help="직접 입력한 보정계수. 생략하면 템플릿 산식을 사용합니다.")
    parser.add_argument("--factor-formula", default=None, help="직접 입력한 보정계수 산식")
    parser.add_argument("--ratio", type=float, default=1.0, help="보정계수에 곱할 비율")
    parser.add_argument("--labor-rates", default=None, help="O/P열 노임단가를 쉼표로 구분해 입력합니다.")
    parser.add_argument("--eval-factor", action="store_true", help="보정계수 산식을 계산하고 종료합니다.")
    parser.add_argument("--keep-templates", action="store_true", help="원본 템플릿 시트를 결과 파일에 남깁니다.")
    parser.add_argument("--gui", action="store_true", help="GUI를 실행합니다.")
    parser.add_argument("--start-watch", action="store_true", help="GUI 실행 후 PDF 폴더 감시를 바로 시작합니다.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    argv = sys.argv[1:] if argv is None else argv
    args = parse_args(argv)
    if args.eval_factor:
        formula = args.factor_formula if args.factor_formula is not None else str(args.factor)
        print(excel_number(evaluate_factor_formula(formula, args.area)))
        return 0
    if args.gui or args.start_watch or not args.pdfs:
        launch_gui(start_watch=args.start_watch)
        return 0

    results = build_workbook(
        template_path=args.template,
        pdf_paths=args.pdfs,
        output_path=args.output,
        area=args.area,
        factor_override=args.factor,
        factor_formula_override=args.factor_formula,
        ratio=args.ratio,
        labor_rates=parse_labor_rates(args.labor_rates),
        keep_templates=args.keep_templates,
    )
    item_count = sum(len(result.items) for result in results)
    print(f"완료: {args.output}")
    print(f"반영 업무 행 수: {item_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
